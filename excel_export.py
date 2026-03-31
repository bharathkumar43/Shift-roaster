import calendar
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from roster_engine import SHIFTS, DAY_ABBR, DAY_NAMES

SHIFT_COLORS = {
    1: PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    2: PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
    3: PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="5D6D7E", end_color="5D6D7E", fill_type="solid")
SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="2C3E50")
OFF_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
OFF_FONT = Font(color="95A5A6", italic=True)
WARNING_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
TAKEOVER_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
TAKEOVER_FONT = Font(color="B7950B", italic=True, size=10)
NO_COVER_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
NO_COVER_FONT = Font(color="C0392B", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_content_types(content_types):
    if isinstance(content_types, list):
        return ", ".join(content_types)
    return str(content_types)


def generate_excel(roster, warnings, shift_assignments, employees, year, month,
                   proj_coverage=None, proj_warnings=None):
    wb = Workbook()

    _create_roster_sheet(wb, roster, shift_assignments, employees, year, month)
    _create_summary_sheet(wb, shift_assignments, employees)
    if proj_coverage:
        _create_project_coverage_sheet(wb, proj_coverage, year, month)
    if warnings or proj_warnings:
        all_warnings = list(warnings or [])
        if proj_warnings:
            all_warnings.append("")
            all_warnings.append("--- PROJECT COVERAGE WARNINGS ---")
            all_warnings.extend(proj_warnings)
        _create_warnings_sheet(wb, all_warnings)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _create_roster_sheet(wb, roster, shift_assignments, employees, year, month):
    ws = wb.active
    ws.title = "Monthly Roster"

    month_name = calendar.month_name[month]
    num_days = calendar.monthrange(year, month)[1]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 1)
    title_cell = ws.cell(row=1, column=1, value=f"Employee Work Load Distribution - {month_name} {year}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_days + 1)
    legend = ws.cell(
        row=2, column=1,
        value="Shift 1: 5AM-2PM IST / 7:30PM-4:30AM EST (Lean)  |  Shift 2: 1PM-10PM IST / 3:30AM-12:30PM EST (Strong)  |  Shift 3: 9PM-6AM IST / 11:30AM-8:30PM EST (Strong)"
    )
    legend.font = Font(size=9, italic=True, color="5D6D7E")
    legend.alignment = Alignment(horizontal="center")

    row = 4
    ws.cell(row=row, column=1, value="Date ->")
    _style_header(ws.cell(row=row, column=1))
    ws.column_dimensions["A"].width = 28

    for day in range(1, num_days + 1):
        d = date(year, month, day)
        weekday = d.weekday()
        col = day + 1
        header_val = f"{day}\n{DAY_ABBR[weekday]}"
        cell = ws.cell(row=row, column=col, value=header_val)
        _style_header(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 14

    row += 1

    for shift_num in [1, 2, 3]:
        shift_info = SHIFTS[shift_num]
        shift_emps = [e for e in employees if shift_assignments[e["name"]] == shift_num]

        if not shift_emps:
            continue

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_days + 1)
        shift_header = ws.cell(
            row=row, column=1,
            value=f"{shift_info['name']} ({shift_info['time_ist']} / {shift_info['time_est']}) - {shift_info['strength'].upper()} team"
        )
        shift_header.font = SUBHEADER_FONT
        shift_header.fill = SUBHEADER_FILL
        shift_header.alignment = Alignment(horizontal="left")
        row += 1

        for emp in shift_emps:
            types_str = _format_content_types(emp["content_types"])
            label = f"{emp['name']} [{types_str}]"
            name_cell = ws.cell(row=row, column=1, value=label)
            name_cell.font = Font(bold=True, size=10)
            name_cell.fill = SHIFT_COLORS[shift_num]
            name_cell.border = THIN_BORDER

            for day in range(1, num_days + 1):
                d = date(year, month, day)
                weekday = d.weekday()
                day_name = DAY_NAMES[weekday]
                col = day + 1
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if day_name in emp["working_days"]:
                    cell.value = "\u2713"
                    cell.fill = SHIFT_COLORS[shift_num]
                    cell.font = Font(color="27AE60", bold=True, size=12)
                else:
                    cell.value = "OFF"
                    cell.fill = OFF_FILL
                    cell.font = OFF_FONT

            row += 1
        row += 1


def _create_summary_sheet(wb, shift_assignments, employees):
    ws = wb.create_sheet("Shift Summary")

    ws.cell(row=1, column=1, value="Shift Distribution Summary")
    ws.cell(row=1, column=1).font = TITLE_FONT

    headers = ["Employee", "Content Types", "Working Days", "Assigned Shift", "Shift Timing"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        _style_header(cell)

    row = 4
    for shift_num in [1, 2, 3]:
        shift_emps = [e for e in employees if shift_assignments[e["name"]] == shift_num]
        for emp in shift_emps:
            ws.cell(row=row, column=1, value=emp["name"]).border = THIN_BORDER
            ws.cell(row=row, column=2,
                    value=_format_content_types(emp["content_types"])).border = THIN_BORDER
            ws.cell(row=row, column=3,
                    value=", ".join(emp["working_days"])).border = THIN_BORDER

            shift_cell = ws.cell(row=row, column=4, value=SHIFTS[shift_num]["name"])
            shift_cell.fill = SHIFT_COLORS[shift_num]
            shift_cell.border = THIN_BORDER

            ws.cell(row=row, column=5,
                    value=f"{SHIFTS[shift_num]['time_ist']} / {SHIFTS[shift_num]['time_est']}").border = THIN_BORDER
            row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25


def _create_project_coverage_sheet(wb, proj_coverage, year, month):
    ws = wb.create_sheet("Project Coverage")

    month_name = calendar.month_name[month]
    ws.cell(row=1, column=1, value=f"Project Coverage - {month_name} {year}")
    ws.cell(row=1, column=1).font = TITLE_FONT

    ws.cell(row=2, column=1,
            value="All 3 shifts shown per project")
    ws.cell(row=2, column=1).font = Font(size=9, italic=True, color="5D6D7E")

    if not proj_coverage or not proj_coverage[0]["projects"]:
        ws.cell(row=4, column=1, value="No projects configured.")
        return

    project_names = []
    seen = set()
    for day_data in proj_coverage:
        for p in day_data["projects"]:
            key = (p["project_name"], p["product_type"])
            if key not in seen:
                seen.add(key)
                project_names.append(key)

    row = 4

    ws.cell(row=row, column=1, value="Date")
    _style_header(ws.cell(row=row, column=1))
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws.column_dimensions["A"].width = 18

    col = 2
    for pname, ptype in project_names:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        cell = ws.cell(row=row, column=col, value=f"{pname} ({ptype})")
        _style_header(cell)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for s_idx, s_label in enumerate(["S1", "S2", "S3"]):
            c = ws.cell(row=row + 1, column=col + s_idx, value=s_label)
            c.font = Font(bold=True, size=9, color="5D6D7E")
            c.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col + s_idx)].width = 16

        col += 3

    row += 2

    for day_data in proj_coverage:
        date_cell = ws.cell(row=row, column=1,
                            value=f"{day_data['date']} ({day_data['day_abbr']})")
        date_cell.font = Font(bold=True, size=10)
        date_cell.border = THIN_BORDER

        proj_map = {}
        for p in day_data["projects"]:
            proj_map[(p["project_name"], p["product_type"])] = p

        col = 2
        for key in project_names:
            p = proj_map.get(key)
            for sn in [1, 2, 3]:
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if not p:
                    cell.value = "-"
                    cell.font = OFF_FONT
                else:
                    sh = p["shifts"].get(sn, {})
                    handler = sh.get("handler")
                    is_secondary = sh.get("is_secondary", False)
                    is_owner_shift = sh.get("is_owner_shift", False)

                    if handler:
                        cell.value = handler
                        cell.font = Font(size=9)
                    else:
                        cell.value = "-"
                        cell.font = OFF_FONT

                col += 1

        row += 1


def _create_warnings_sheet(wb, warnings):
    ws = wb.create_sheet("Coverage Warnings")

    ws.cell(row=1, column=1, value="Coverage Gap Warnings")
    ws.cell(row=1, column=1).font = TITLE_FONT

    cell = ws.cell(row=3, column=1, value="Warning")
    _style_header(cell)
    ws.column_dimensions["A"].width = 80

    for i, w in enumerate(warnings, 4):
        cell = ws.cell(row=i, column=1, value=w)
        if w and not w.startswith("---"):
            cell.fill = WARNING_FILL
        cell.border = THIN_BORDER


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
