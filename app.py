import calendar
import json
import os
from datetime import date
from functools import wraps
from io import BytesIO

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

from flask import (Flask, render_template, request, session, g,
                   redirect, url_for, send_file, flash, jsonify)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

import msal

from roster_engine import (
    generate_roster,
    generate_roster_from_manual_assignments,
    get_roster_summary,
    SHIFTS,
    DAY_NAMES,
    CONTENT_TYPES,
    normalize_five_day_pattern,
    prepare_employees_for_roster_month,
    is_emp_scheduled_work_day,
    is_pinned_shift_1,
    pattern_for_calendar_month,
    coerce_to_five_day_pattern,
)
from excel_export import generate_excel, generate_delta_image
from project_engine import generate_project_coverage
from file_parser import parse_file
import database as db

AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID", "")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET", "")
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID", "")
AZURE_AUTHORITY = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}" if AZURE_TENANT_ID else ""
AZURE_SCOPE = ["User.Read"]
AZURE_ENABLED = bool(AZURE_CLIENT_ID and AZURE_CLIENT_SECRET and AZURE_TENANT_ID)
# Explicit redirect URI so it matches exactly what is registered in Azure portal.
# Locally set to http://localhost:5050/login/azure/callback
# In production set to https://roster.cftools.live/login/azure/callback
AZURE_REDIRECT_URI = os.getenv("AZURE_REDIRECT_URI", "")

# Maps each CloudFuze Microsoft email (lowercase) to the legacy short username
# that person used when they first signed up via password login.
# On their first Microsoft login the username is automatically upgraded to the email.
_LEGACY_USERNAME_MAP = {
    "abhishek.sakala@cloudfuze.com":        "abhishek",
    "abhishikth.yenugula@cloudfuze.com":    "abhishikth",
    "ajay.singh@cloudfuze.com":             "ajay",
    "ramana.reddy@cloudfuze.com":           "ramana",
    "amulya.anapuram@cloudfuze.com":        "amulya",
    "arun@cloudfuze.com":                   "arun",
    "bharath.tummaganti@cloudfuze.com":     "bharath",
    "chaitanya.gupta@cloudfuze.com":        "chaitanya",
    "chandra.mouli@cloudfuze.com":          "chandra mouli",
    "dathu.kaluvala@cloudfuze.com":         "dathu",
    "davidraj.dumpala@cloudfuze.com":       "davidraj",
    "ganesh.kondameedi@cloudfuze.com":      "ganesh",
    "habeebunnisa.begum@cloudfuze.com":     "habeebunnisa",
    "harika.velidi@cloudfuze.com":          "harika",
    "harshith.kaduluri@cloudfuze.com":      "harshith",
    "jyoshitha.dhannapaneni@cloudfuze.com": "jyoshitha",
    "lakshmareddy@cloudfuze.com":           "lakshma reddy",
    "lakshmi.prasanna@cloudfuze.com":       "lakshmi",
    "manoj.bathula@cloudfuze.com":          "manoj",
    "meena.lakshmi@cloudfuze.com":          "meena",
    "neelima.krotta@cloudfuze.com":         "neelima",
    "pallavi.kosuvaripalli@cloudfuze.com":  "pallavi",
    "paresh.rahate@cloudfuze.com":          "paresh",
    "pranavi@cloudfuze.com":                "pranavi",
    "raghu.yellani@cloudfuze.com":          "raghu",
    "ranadeep.muddam@cloudfuze.com":        "ranadeep",
    "ravi.hemanth@cloudfuze.com":           "ravi",
    "saikumar.kustapuram@cloudfuze.com":    "sai",
    "siva.kota@cloudfuze.com":              "siva",
    "sravan.kesaram@cloudfuze.com":         "sravan",
    "sriram.ramakrishnan@cloudfuze.com":    "sriram",
    "subhakar.dhanireddy@cloudfuze.com":    "subhakar",
    "swaroop@cloudfuze.com":                "swaroop",
    "vijendar.burgula@cloudfuze.com":       "vijendar",
    "vineetha.yenti@cloudfuze.com":         "vineetha",
}


def _build_msal_app():
    return msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=AZURE_AUTHORITY,
        client_credential=AZURE_CLIENT_SECRET,
    )


def _get_leave_dates(year, month):
    """Helper to fetch leave dates map for coverage calculations."""
    return db.get_leave_dates_map(year, month)


def _calendar_prev_month(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def _weekoff_roster_rows(employees, year, month):
    """
    Week-off pairs for the roster month vs the prior calendar month, for roster UI.
    employees: list of engineer dicts from DB (with monthly_working_days).
    """
    py, pm = _calendar_prev_month(year, month)
    rows = []
    for e in employees:
        pat_cur = coerce_to_five_day_pattern(pattern_for_calendar_month(e, year, month))
        pat_prev = coerce_to_five_day_pattern(pattern_for_calendar_month(e, py, pm))
        wo_cur = [d for d in DAY_NAMES if d not in pat_cur]
        wo_prev = [d for d in DAY_NAMES if d not in pat_prev]
        rows.append(
            {
                "name": e["name"],
                "working_days": pat_cur,
                "week_offs": wo_cur,
                "prev_week_offs": wo_prev,
                "weekoffs_changed": wo_cur != wo_prev,
            }
        )
    rows.sort(key=lambda r: r["name"].lower())
    return rows


def _roster_weekoff_context(employees, year, month):
    py, pm = _calendar_prev_month(year, month)
    return {
        "weekoff_rows": _weekoff_roster_rows(employees, year, month),
        "weekoff_prev_month_name": calendar.month_name[pm],
        "weekoff_prev_year": py,
    }


app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = os.getenv("SECRET_KEY", "roster-automation-secret-key-change-in-production")

APP_NAME = "Employee Work Load Distribution"

db.init_db()

# Session keys: which calendar month the signed-in user is "working in" across the UI.
SESSION_APP_YEAR = "app_context_year"
SESSION_APP_MONTH = "app_context_month"


def get_app_context_ym():
    """Calendar month/year used for roster defaults, search, summary, leaves, etc."""
    today = date.today()
    try:
        y = int(session.get(SESSION_APP_YEAR, today.year))
        m = int(session.get(SESSION_APP_MONTH, today.month))
    except (TypeError, ValueError):
        return today.year, today.month
    if m < 1 or m > 12 or y < 2000 or y > 2100:
        return today.year, today.month
    return y, m


# ── Auth helpers ─────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.is_json:
                return jsonify({"error": "Authentication required"}), 401
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login", next=request.path))
        if not g.user or g.user.get("role") != "admin":
            if request.is_json:
                return jsonify({"error": "Admin access required"}), 403
            flash("You don't have permission to perform this action.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    g.user = db.get_user_by_id(user_id) if user_id else None


@app.context_processor
def inject_user():
    linked_emp = None
    if g.user and g.user.get("employee_id"):
        linked_emp = db.get_employee_by_id(g.user["employee_id"])
    cy, cm = get_app_context_ym()
    calendar_month_names = [calendar.month_name[m] for m in range(1, 13)]
    return {
        "current_user": g.user,
        "linked_employee": linked_emp,
        "app_context_year": cy,
        "app_context_month": cm,
        "app_context_month_name": calendar.month_name[cm],
        "calendar_month_names": calendar_month_names,
    }


def _predefined(year, month):
    """
    Fixed Shift-1 assignments for roster generation only (pinned engineers).
    """
    merged = {}
    for e in db.get_employees_by_role("engineer"):
        if is_pinned_shift_1(e["name"]):
            merged[e["name"]] = 1
    return merged if merged else None


def _predefined_with_saved_shifts(year, month):
    """
    Engineer fixed shifts for roster generation: Shift-1 pins plus any assignments
    stored in rotation_history for this month (Edit Shifts). Other engineers are
    filled by the engine to meet mandated shift sizes.
    """
    fixed = dict(_predefined(year, month) or {})
    saved = db.get_shift_assignments_for_month(year, month) or {}
    eng_names = {e["name"] for e in db.get_employees_by_role("engineer")}
    for name, shift in saved.items():
        if name in eng_names:
            fixed[name] = int(shift)
    for e in db.get_employees_by_role("engineer"):
        if is_pinned_shift_1(e["name"]):
            fixed[e["name"]] = 1
    return fixed if fixed else None


def _generate_roster_with_saved_month(
    employees, year, month, night_counts=None, *, strict=False
):
    """
    Run roster generation honoring saved month shifts.

    If admin has assigned ALL engineers via Edit Shifts, those assignments are used
    directly with no cap enforcement. Otherwise auto-distributes unassigned engineers.
    """
    nc = night_counts if night_counts is not None else db.get_night_shift_counts()
    saved = db.get_shift_assignments_for_month(year, month) or {}

    # Build complete map: pins override, then saved Edit Shifts assignments
    complete = {}
    for e in employees:
        nm = e["name"]
        if is_pinned_shift_1(nm):
            complete[nm] = 1
        elif nm in saved:
            complete[nm] = int(saved[nm])

    # Admin has assigned every engineer — use those directly, no cap enforcement
    if len(complete) == len(employees) and complete:
        return generate_roster_from_manual_assignments(
            employees,
            year,
            month,
            complete,
            prev_month_night_ids=_prev_month_night_ids(year, month),
        )

    # Partial or no assignments — auto-distribute remaining engineers
    pre = _predefined_with_saved_shifts(year, month)
    try:
        return _generate_roster(employees, year, month, nc, pre)
    except ValueError:
        if strict:
            raise
        if len(complete) < len(employees):
            _, _, guess = _generate_roster(employees, year, month, nc, _predefined(year, month))
            for e in employees:
                nm = e["name"]
                if nm not in complete:
                    complete[nm] = int(guess.get(nm, 1))
        return generate_roster_from_manual_assignments(
            employees,
            year,
            month,
            complete,
            prev_month_night_ids=_prev_month_night_ids(year, month),
        )


def _prev_month_night_ids(year, month):
    """Employee ids on Shift 3 (night) in the previous calendar month."""
    if month == 1:
        py, pm = year - 1, 12
    else:
        py, pm = year, month - 1
    return frozenset(db.get_employee_ids_on_shift(py, pm, 3))


def _complete_engineer_shifts_for_save(engineers, assignments, year, month):
    """
    One shift number per engineer for manual roster build: form body, pins, then DB,
    then a single auto guess for anyone still missing (unusual).
    """
    saved = dict(db.get_shift_assignments_for_month(year, month) or {})
    complete = {}
    for e in engineers:
        nm = e["name"]
        if is_pinned_shift_1(nm):
            complete[nm] = 1
        elif nm in assignments:
            complete[nm] = int(assignments[nm])
        elif nm in saved:
            complete[nm] = int(saved[nm])
    if len(complete) < len(engineers):
        # Avoid assign_shifts here (can reject DB/saved ratios); fill gaps with rotation.
        idx = 0
        for e in engineers:
            nm = e["name"]
            if nm not in complete:
                complete[nm] = (idx % 3) + 1
                idx += 1
    return complete


def _generate_roster(
    employees, year, month, night_counts=None, predefined=None, *, relax_fixed_caps=False
):
    """generate_roster with night-shift anti-streak (no back-to-back months on Shift 3)."""
    nc = night_counts if night_counts is not None else db.get_night_shift_counts()
    pre = predefined if predefined is not None else _predefined(year, month)
    return generate_roster(
        employees,
        year,
        month,
        nc,
        pre,
        prev_month_night_ids=_prev_month_night_ids(year, month),
        relax_fixed_caps=relax_fixed_caps,
    )


@app.route("/set_app_context", methods=["GET", "POST"])
@admin_required
def set_app_context():
    """Persist selected calendar month for roster/search/summary/leaves defaults."""
    if request.method == "GET":
        return redirect(url_for("index"))
    try:
        year = int(request.form.get("year", 0))
        month = int(request.form.get("month", 0))
    except (TypeError, ValueError):
        flash("Invalid month or year.", "danger")
        return redirect(url_for("index"))
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        flash("Invalid month or year.", "danger")
        return redirect(url_for("index"))
    session[SESSION_APP_YEAR] = year
    session[SESSION_APP_MONTH] = month
    flash(f"Application month set to {calendar.month_name[month]} {year}.", "success")
    nxt = (request.form.get("next") or "").strip()
    if not nxt.startswith("/"):
        nxt = url_for("index")
    return redirect(nxt)


# ── Index / Employee Management ──────────────────────────

@app.route("/", methods=["GET"])
@login_required
def index():
    engineers = db.get_employees_by_role("engineer")
    shift_leads = db.get_employees_by_role("shift_lead")
    managers = db.get_employees_by_role("manager")
    all_projects = db.get_all_projects()
    saved_rosters = db.list_saved_rosters()
    finalized_rosters = db.list_finalized_rosters()
    today = date.today()
    this_month = today.month
    this_year = today.year
    ctx_y, ctx_m = get_app_context_ym()
    default_roster_month = ctx_m
    default_roster_year = ctx_y
    month_names = [calendar.month_name[m] for m in range(1, 13)]
    imported = request.args.get("imported", type=int)
    if imported is not None:
        flash(f"Successfully imported {imported} employee(s).", "success")

    shift_assignments = db.get_shift_assignments_for_month(ctx_y, ctx_m)

    engineers_display = []
    for e in engineers:
        row = dict(e)
        pat = coerce_to_five_day_pattern(pattern_for_calendar_month(e, ctx_y, ctx_m))
        row["generated_working_days"] = pat
        row["generated_week_offs"] = [d for d in DAY_NAMES if d not in pat]
        row["shift"] = shift_assignments.get(e["name"])
        engineers_display.append(row)

    return render_template("index.html",
                           app_name=APP_NAME,
                           engineers=engineers_display,
                           shift_leads=shift_leads,
                           managers=managers,
                           employees=engineers_display,
                           all_projects=all_projects,
                           saved_rosters=saved_rosters,
                           finalized_rosters=finalized_rosters,
                           day_names=DAY_NAMES,
                           content_types=CONTENT_TYPES,
                           this_month=this_month,
                           this_year=this_year,
                           default_roster_month=default_roster_month,
                           default_roster_year=default_roster_year,
                           default_search_month=ctx_m,
                           default_search_year=ctx_y,
                           current_year=this_year,
                           year_range_lo=min(this_year - 1, ctx_y - 1),
                           year_range_hi=max(this_year + 3, ctx_y + 2),
                           month_names=month_names)


ROLE_LABELS = {
    "engineer": "Migration Engineer",
    "shift_lead": "Shift Lead",
    "manager": "Migration Manager",
}


@app.route("/clear_all_saved_rosters", methods=["POST"])
@admin_required
def clear_all_saved_rosters():
    """Wipe saved roster months, rotation history, and monthly week-off snapshots."""
    db.clear_all_saved_rosters_and_rotation()
    session.pop("draft_roster", None)
    session.pop("last_roster", None)
    flash(
        "All saved rosters, shift rotation history, and stored monthly week-off snapshots "
        "were removed. Generate and save a new roster when ready.",
        "success",
    )
    return redirect(url_for("index"))


@app.route("/add_employee", methods=["POST"])
@admin_required
def add_employee():
    name = request.form.get("name", "").strip()
    content_types = request.form.getlist("content_types")
    working_days = request.form.getlist("working_days")
    emp_role = request.form.get("emp_role", "engineer")
    project_names = request.form.getlist("project_name")
    project_types = request.form.getlist("project_type")

    if emp_role not in ROLE_LABELS:
        emp_role = "engineer"

    if emp_role in ("manager", "shift_lead"):
        if not content_types:
            content_types = ["Content"]
        if not working_days:
            working_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    elif not name or not content_types or len(working_days) == 0:
        flash("Please fill in all required fields.", "danger")
        return redirect(url_for("index"))

    try:
        working_days = normalize_five_day_pattern(working_days)
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("index"))

    if not name:
        flash("Please enter a name.", "danger")
        return redirect(url_for("index"))

    emp_id = db.add_employee(name, content_types, working_days, emp_role=emp_role)
    if emp_id is None:
        flash(f"'{name}' already exists.", "warning")
        return redirect(url_for("index"))

    for pname, ptype in zip(project_names, project_types):
        pname = pname.strip()
        if pname and ptype:
            db.add_project(pname, ptype, emp_id)

    label = ROLE_LABELS.get(emp_role, "Employee")
    flash(f"{label} '{name}' added successfully.", "success")
    return redirect(url_for("index"))


@app.route("/edit_employee/<int:emp_id>", methods=["POST"])
@admin_required
def edit_employee(emp_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid data"}), 400
    new_name = (data.get("name") or "").strip() or None
    content_types = data.get("content_types") or ["Content"]
    working_days = data.get("working_days") or ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    try:
        working_days = normalize_five_day_pattern(working_days)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.update_employee(emp_id, content_types, working_days, name=new_name)
    # Keep app-month week-offs aligned with edits so roster updates immediately.
    emp_row = db.get_employee_by_id(emp_id)
    if emp_row and emp_row.get("emp_role") == "engineer":
        cy, cm = get_app_context_ym()
        db.snapshot_monthly_working_pattern(emp_id, cy, cm, working_days)
    # Replace projects
    db.clear_projects_for_employee(emp_id)
    for proj in (data.get("projects") or []):
        name = (proj.get("name") or "").strip()
        ptype = proj.get("product_type") or "Content"
        if name:
            db.add_project(name, ptype, emp_id)
    return jsonify({"success": True})


@app.route("/remove_employee/<int:emp_id>", methods=["POST"])
@admin_required
def remove_employee(emp_id):
    db.remove_employee(emp_id)
    return redirect(url_for("index"))


@app.route("/clear_all", methods=["POST"])
@admin_required
def clear_all():
    db.clear_all_employees()
    flash("All employees cleared.", "success")
    return redirect(url_for("index"))


# ── Roster Generation ────────────────────────────────────

def _build_roster_data(employees, roster, shift_assignments, year, month):
    """Build the roster_data list used by the template."""
    num_days = calendar.monthrange(year, month)[1]
    roster_data = []
    for day in range(1, num_days + 1):
        d = date(year, month, day)
        weekday = d.weekday()
        day_info = {
            "date": d.strftime("%b %d"),
            "day_name": DAY_NAMES[weekday],
            "day_abbr": d.strftime("%a"),
            "day_num": day,
            "shifts": {}
        }
        for shift_num in [1, 2, 3]:
            emp_names = roster[d][shift_num]
            emp_details = []
            for name in emp_names:
                emp = next(e for e in employees if e["name"] == name)
                emp_details.append({
                    "name": name,
                    "content_types": emp["content_types"]
                })
            day_info["shifts"][shift_num] = emp_details
        roster_data.append(day_info)
    return roster_data


def _excel_bytes_from_shift_assignments(year, month, shift_assignments, warnings=None):
    """
    Build an Excel workbook from engineer shift numbers (same engine as the roster grid).

    Used for downloads so the file matches the current draft or saved roster JSON, not a
    stale excel_blob left over from an earlier save.
    """
    employees = db.get_employees_by_role("engineer")
    night_counts = db.get_night_shift_counts()
    if len(employees) < 2:
        roster, gen_warnings, sa = _generate_roster_with_saved_month(
            employees, year, month, night_counts
        )
    else:
        eng_names = {e["name"] for e in employees}
        pre = {}
        for name, shift in (shift_assignments or {}).items():
            if name not in eng_names:
                continue
            try:
                pre[name] = int(shift)
            except (TypeError, ValueError):
                continue
        for e in employees:
            if is_pinned_shift_1(e["name"]):
                pre[e["name"]] = 1
        try:
            roster, gen_warnings, sa = _generate_roster(
                employees, year, month, night_counts, pre or None
            )
        except ValueError:
            complete = {}
            raw = shift_assignments or {}
            for e in employees:
                nm = e["name"]
                if is_pinned_shift_1(nm):
                    complete[nm] = 1
                elif nm in pre:
                    complete[nm] = pre[nm]
                else:
                    try:
                        complete[nm] = int(raw.get(nm, raw.get(str(nm), 2)))
                    except (TypeError, ValueError):
                        complete[nm] = 2
            try:
                roster, gen_warnings, sa = generate_roster_from_manual_assignments(
                    employees,
                    year,
                    month,
                    complete,
                    prev_month_night_ids=_prev_month_night_ids(year, month),
                )
            except (ValueError, KeyError):
                roster, gen_warnings, sa = _generate_roster(
                    employees,
                    year,
                    month,
                    night_counts,
                    predefined=complete,
                    relax_fixed_caps=True,
                )
    merged_warnings = list(warnings or []) + list(gen_warnings or [])
    all_projects = db.get_all_projects()
    proj_coverage, proj_warnings = generate_project_coverage(
        all_projects, employees, sa, year, month, _get_leave_dates(year, month)
    )
    return generate_excel(
        roster,
        merged_warnings,
        sa,
        employees,
        year,
        month,
        proj_coverage,
        proj_warnings,
    )


@app.route("/generate", methods=["GET", "POST", "OPTIONS"])
@login_required
def generate():
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "GET":
        flash(
            "Roster generation runs from the home page: pick month/year, then click Shift Roster. "
            "Opening /generate directly in the browser is not supported.",
            "info",
        )
        return redirect(url_for("index"))
    employees = db.get_employees_by_role("engineer")
    if len(employees) < 2:
        flash("Add at least 2 employees to generate a roster.", "warning")
        return redirect(url_for("index"))

    cy, cm = get_app_context_ym()
    year = int(request.form.get("year", cy))
    month = int(request.form.get("month", cm))

    saved = db.get_saved_roster_data(year, month)
    if saved and not request.form.get("regenerate"):
        return _render_saved_roster(year, month, saved)

    # Keep this month's snapshot so explicit week-off edits are honored.
    employees = db.get_employees_by_role("engineer")

    night_counts = db.get_night_shift_counts()
    try:
        roster, warnings, shift_assignments = _generate_roster_with_saved_month(
            employees, year, month, night_counts, strict=True
        )
    except ValueError as err:
        flash(
            f"Cannot generate roster: {err} "
            "Adjust or clear manual shift assignments for this month (Edit Shifts), then try again.",
            "danger",
        )
        return redirect(url_for("index"))

    # Persist this month's working pattern so later months can rotate week-offs off it
    # even if the user has not clicked Save yet (Save still re-writes the same keys).
    prepared, prep_warnings = prepare_employees_for_roster_month(employees, year, month)
    for pe in prepared:
        db.snapshot_monthly_working_pattern(pe["id"], year, month, pe["working_days"])
    employees = db.get_employees_by_role("engineer")
    if prep_warnings:
        warnings = list(warnings or []) + list(prep_warnings)

    summary = get_roster_summary(employees, shift_assignments)
    month_name = calendar.month_name[month]
    roster_data = _build_roster_data(employees, roster, shift_assignments, year, month)
    wo_ctx = _roster_weekoff_context(employees, year, month)

    session["last_roster"] = {"year": year, "month": month}
    session["draft_roster"] = {
        "year": year,
        "month": month,
        "roster_data": roster_data,
        "warnings": warnings,
        "summary": {str(k): v for k, v in summary.items()},
        "shift_assignments": shift_assignments,
    }

    return render_template(
        "roster.html",
        app_name=APP_NAME,
        roster_data=roster_data,
        warnings=warnings,
        summary=summary,
        shifts=SHIFTS,
        month_name=month_name,
        year=year,
        month=month,
        is_draft=True,
        is_saved=False,
        **wo_ctx,
    )


@app.route("/save_roster", methods=["POST"])
@login_required
def save_roster():
    draft = session.get("draft_roster")
    if not draft:
        flash("No draft roster to save.", "warning")
        return redirect(url_for("index"))

    year = draft["year"]
    month = draft["month"]

    employees = db.get_employees_by_role("engineer")
    shift_assignments = draft["shift_assignments"]

    db.save_all_rotations(shift_assignments, employees, year, month)

    employees = db.get_employees_by_role("engineer")

    prepared, _ = prepare_employees_for_roster_month(employees, year, month)
    for pe in prepared:
        db.snapshot_monthly_working_pattern(pe["id"], year, month, pe["working_days"])

    all_projects = db.get_all_projects()
    _excl = db.get_excluded_projects()
    all_projects = [p for p in all_projects if (p["name"], p["product_type"]) not in _excl]
    night_counts = db.get_night_shift_counts()
    try:
        roster, _, _ = _generate_roster_with_saved_month(
            employees, year, month, night_counts
        )
    except Exception:
        roster = []
    try:
        proj_coverage, proj_warnings = generate_project_coverage(
            all_projects, employees, shift_assignments, year, month, _get_leave_dates(year, month)
        )
    except Exception:
        proj_coverage, proj_warnings = [], []

    excel_output = generate_excel(
        roster, draft["warnings"], shift_assignments, employees, year, month,
        proj_coverage, proj_warnings
    )
    db.save_roster_excel(year, month, excel_output.read())

    roster_json = json.dumps({
        "roster_data": draft["roster_data"],
        "warnings": draft["warnings"],
        "summary": draft["summary"],
        "shift_assignments": shift_assignments,
    })
    db.save_roster_data(year, month, roster_json)

    session.pop("draft_roster", None)

    month_name = calendar.month_name[month]
    flash(f"Roster for {month_name} {year} has been saved.", "success")
    return redirect(url_for("view_roster", year=year, month=month))


@app.route("/view_roster")
@login_required
def view_roster():
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        flash("Please specify a month and year.", "warning")
        return redirect(url_for("index"))

    saved = db.get_saved_roster_data(year, month)
    if not saved:
        flash(f"No saved roster found for {calendar.month_name[month]} {year}.", "warning")
        return redirect(url_for("index"))

    return _render_saved_roster(year, month, saved)


def _render_saved_roster(year, month, saved):
    data = json.loads(saved["roster_json"])
    summary_raw = data["summary"]
    summary = {int(k): v for k, v in summary_raw.items()}
    month_name = calendar.month_name[month]
    engineers = db.get_employees_by_role("engineer")
    wo_ctx = _roster_weekoff_context(engineers, year, month)

    session["last_roster"] = {"year": year, "month": month}

    return render_template(
        "roster.html",
        app_name=APP_NAME,
        roster_data=data["roster_data"],
        warnings=data["warnings"],
        summary=summary,
        shifts=SHIFTS,
        month_name=month_name,
        year=year,
        month=month,
        is_draft=False,
        is_saved=True,
        saved_at=saved["saved_at"],
        **wo_ctx,
    )


# ── Project Coverage ─────────────────────────────────────

@app.route("/projects", methods=["GET", "POST"])
@login_required
def projects():
    employees = db.get_employees_by_role("engineer")
    all_projects = db.get_all_projects()

    if not employees or not all_projects:
        flash("Add employees and projects first.", "warning")
        return redirect(url_for("index"))

    cy, cm = get_app_context_ym()
    year = int(request.form.get("year", request.args.get("year", cy)))
    month = int(request.form.get("month", request.args.get("month", cm)))

    night_counts = db.get_night_shift_counts()
    try:
        _, _, shift_assignments = _generate_roster_with_saved_month(
            employees, year, month, night_counts
        )
    except Exception:
        flash("Roster not yet generated for this month. Please generate and save the roster first.", "warning")
        return redirect(url_for("index"))

    try:
        coverage, proj_warnings = generate_project_coverage(
            all_projects, employees, shift_assignments, year, month, _get_leave_dates(year, month)
        )
    except Exception as e:
        app.logger.exception("generate_project_coverage failed")
        if g.user and g.user.get("role") == "admin":
            flash(f"Coverage error: {e}", "danger")
        coverage, proj_warnings = [], []

    month_name = calendar.month_name[month]
    session["last_roster"] = {"year": year, "month": month}

    all_people = db.get_all_employees()
    emp_role_map = {e["name"]: e.get("emp_role", "engineer") for e in all_people}
    engineer_projects = [p for p in all_projects if emp_role_map.get(p["employee_name"]) == "engineer"]
    lead_projects = [p for p in all_projects if emp_role_map.get(p["employee_name"]) == "shift_lead"]
    manager_projects = [p for p in all_projects if emp_role_map.get(p["employee_name"]) == "manager"]

    proj_manager_map = {}
    for mp in manager_projects:
        proj_manager_map[mp["name"]] = mp["employee_name"]

    return render_template("projects.html",
                           app_name=APP_NAME,
                           coverage=coverage,
                           warnings=proj_warnings,
                           month_name=month_name,
                           year=year,
                           month=month,
                           projects=all_projects,
                           engineer_projects=engineer_projects,
                           lead_projects=lead_projects,
                           manager_projects=manager_projects,
                           proj_manager_map=proj_manager_map,
                           shifts=SHIFTS,
                           shift_assignments=shift_assignments)


# ── Download Excel ───────────────────────────────────────

@app.route("/download", methods=["GET"])
@login_required
def download():
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    if not year or not month:
        last = session.get("last_roster", {})
        cy, cm = get_app_context_ym()
        year = last.get("year", cy)
        month = last.get("month", cm)

    month_name = calendar.month_name[month]
    filename = f"Workload_{month_name}_{year}.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    from_draft = str(request.args.get("draft") or "").strip().lower() in ("1", "true", "yes")
    draft = session.get("draft_roster")
    draft_matches = bool(draft and draft.get("year") == year and draft.get("month") == month)
    saved_row = db.get_saved_roster_data(year, month)

    def _send_excel_from_draft():
        raw_sa = draft.get("shift_assignments") or {}
        sa = {}
        for k, v in raw_sa.items():
            try:
                sa[str(k)] = int(v)
            except (TypeError, ValueError):
                continue
        output = _excel_bytes_from_shift_assignments(
            year, month, sa, warnings=draft.get("warnings")
        )
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype,
        )

    # Explicit draft download (roster page with unsaved changes / current generate).
    if from_draft and draft_matches:
        return _send_excel_from_draft()

    # Finalized month — rebuild from saved JSON (matches screen; fixes stale excel_blob).
    if saved_row:
        data = json.loads(saved_row["roster_json"])
        raw_sa = data.get("shift_assignments") or {}
        sa = {}
        for k, v in raw_sa.items():
            try:
                sa[str(k)] = int(v)
            except (TypeError, ValueError):
                continue
        output = _excel_bytes_from_shift_assignments(
            year, month, sa, warnings=data.get("warnings")
        )
        excel_bytes = output.read()
        db.save_roster_excel(year, month, excel_bytes)
        return send_file(
            BytesIO(excel_bytes),
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype,
        )

    # Draft in session but no saved JSON for this month (e.g. old links without ?draft=1).
    if draft_matches:
        return _send_excel_from_draft()

    # Legacy: excel stored without finalized JSON (unusual).
    saved_blob = db.get_saved_roster(year, month)
    if saved_blob:
        return send_file(
            BytesIO(saved_blob),
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype,
        )

    employees = db.get_employees_by_role("engineer")
    night_counts = db.get_night_shift_counts()
    try:
        roster, warnings, shift_assignments = _generate_roster_with_saved_month(
            employees, year, month, night_counts
        )
    except Exception:
        flash("Roster not yet generated for this month. Please generate and save the roster first.", "warning")
        return redirect(url_for("index"))

    all_projects = db.get_all_projects()
    try:
        proj_coverage, proj_warnings = generate_project_coverage(
            all_projects, employees, shift_assignments, year, month, _get_leave_dates(year, month)
        )
    except Exception:
        proj_coverage, proj_warnings = [], []

    output = generate_excel(
        roster, warnings, shift_assignments, employees, year, month,
        proj_coverage, proj_warnings
    )

    excel_tail = output.read()
    db.save_roster_excel(year, month, excel_tail)
    return send_file(
        BytesIO(excel_tail),
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype,
    )


# ── File Upload ──────────────────────────────────────────

@app.route("/preview_upload", methods=["POST"])
@admin_required
def preview_upload():
    """Parse uploaded file and return employee list as JSON for the preview/edit modal."""
    if "file" not in request.files:
        return jsonify({"error": "No file selected"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400
    try:
        employees = parse_file(file)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"employees": employees})


@app.route("/upload_confirm", methods=["POST"])
@admin_required
def upload_confirm():
    """Receive the (possibly edited) employee list as JSON and save to DB."""
    data = request.get_json()
    if not data or "employees" not in data:
        return jsonify({"error": "Invalid data"}), 400
    employees = data["employees"]
    if not employees:
        return jsonify({"error": "No employees provided"}), 400
    db.clear_all_employees()
    count = 0
    for emp in employees:
        name = (emp.get("name") or "").strip()
        if not name:
            continue
        content_types = emp.get("content_types") or ["Content"]
        working_days = emp.get("working_days") or ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        try:
            working_days = normalize_five_day_pattern(working_days)
        except ValueError as e:
            return jsonify({"error": f"{name}: {e}"}), 400
        emp_id = db.add_employee(name, content_types, working_days)
        if emp_id:
            count += 1
            for proj in (emp.get("projects") or []):
                proj_name = (proj.get("name") or "").strip()
                proj_type = proj.get("product_type") or "Content"
                if proj_name:
                    db.add_project(proj_name, proj_type, emp_id)
    return jsonify({"success": True, "count": count})


@app.route("/upload", methods=["POST"])
@admin_required
def upload():
    if "file" not in request.files:
        flash("No file selected.", "danger")
        return redirect(url_for("index"))

    file = request.files["file"]
    if not file.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("index"))

    try:
        employees = parse_file(file)
    except ValueError as e:
        flash(f"Upload failed: {e}", "danger")
        return redirect(url_for("index"))
    except Exception as e:
        flash(f"Error processing file: {e}", "danger")
        return redirect(url_for("index"))

    db.clear_all_employees()

    count = 0
    for emp in employees:
        try:
            wd = normalize_five_day_pattern(emp["working_days"])
        except ValueError as e:
            flash(f"{emp.get('name', '?')}: {e}", "danger")
            return redirect(url_for("index"))
        emp_id = db.add_employee(emp["name"], emp["content_types"], wd)
        if emp_id:
            count += 1

    flash(
        f"Successfully imported {count} employees from '{file.filename}'. "
        f"Shifts will be redistributed when you generate a roster.",
        "success"
    )
    return redirect(url_for("index"))


# ── Search API ───────────────────────────────────────────

@app.route("/search", methods=["GET"])
@login_required
def search():
    q = request.args.get("q", "").strip()
    cy, cm = get_app_context_ym()
    year = request.args.get("year", cy, type=int)
    month = request.args.get("month", cm, type=int)

    if not q or len(q) < 1:
        return jsonify({"employees": [], "projects": []})

    employees = db.get_employees_by_role("engineer")
    night_counts = db.get_night_shift_counts()

    shift_assignments = {}
    prepared_lookup = {}
    if employees:
        try:
            _, _, shift_assignments = _generate_roster_with_saved_month(
                employees, year, month, night_counts
            )
            prepared, _ = prepare_employees_for_roster_month(employees, year, month)
            prepared_lookup = {e["name"]: e for e in prepared}
        except Exception:
            pass

    emp_results = []
    matched_emps = db.search_employees(q)
    all_projects = db.get_all_projects()

    for emp in matched_emps:
        emp_projects = [p for p in all_projects if p["employee_id"] == emp["id"]]
        shift = shift_assignments.get(emp["name"])
        shift_info = None
        if shift:
            shift_info = {
                "number": shift,
                "name": SHIFTS[shift]["name"],
                "time_ist": SHIFTS[shift]["time_ist"],
                "time_est": SHIFTS[shift]["time_est"],
                "strength": SHIFTS[shift]["strength"],
            }

        num_days = calendar.monthrange(year, month)[1]
        working_count = 0
        off_count = 0
        sched = prepared_lookup.get(emp["name"])
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            day_name = DAY_NAMES[d.weekday()]
            if sched:
                on = is_emp_scheduled_work_day(sched, d)
            else:
                on = day_name in (emp.get("working_days") or DAY_NAMES[:5])
            if on:
                working_count += 1
            else:
                off_count += 1

        emp_results.append({
            "name": emp["name"],
            "content_types": emp["content_types"],
            "working_days": emp["working_days"],
            "projects": [{"name": p["name"], "product_type": p["product_type"]}
                         for p in emp_projects],
            "shift": shift_info,
            "working_days_count": working_count,
            "off_days_count": off_count,
        })

    all_projects = db.get_all_projects()
    proj_coverage = []
    if employees and shift_assignments:
        try:
            proj_coverage_data, _ = generate_project_coverage(
                all_projects, employees, shift_assignments, year, month, _get_leave_dates(year, month)
            )
            proj_coverage = proj_coverage_data
        except Exception:
            pass

    all_managers = db.get_employees_by_role("manager")
    all_leads = db.get_employees_by_role("shift_lead")
    all_people = db.get_all_employees()

    proj_results = []
    seen_projects = set()
    matched_projs = db.search_projects(q)

    emp_lookup = {e["name"]: True for e in employees}

    for proj in matched_projs:
        is_engineer_proj = proj["employee_name"] in emp_lookup
        proj_key = (proj["name"], proj["product_type"], proj["employee_name"] if is_engineer_proj else "_mgr_")

        name_type_key = (proj["name"], proj["product_type"])
        if name_type_key in seen_projects:
            continue
        if not is_engineer_proj:
            if any(p["name"] == proj["name"] and p["product_type"] == proj["product_type"] and p["employee_name"] in emp_lookup for p in matched_projs):
                continue
        seen_projects.add(name_type_key)

        same_name_projs = [p for p in all_projects if p["name"] == proj["name"]]
        manager_name = None
        for sp in same_name_projs:
            mgr = next((m for m in all_managers if m["id"] == sp["employee_id"]), None)
            if mgr:
                manager_name = mgr["name"]
                break

        engineer_name = None
        shift = None
        shift_name = None
        for day_data in proj_coverage:
            for p in day_data.get("projects", []):
                if p["project_name"].lower() == proj["name"].lower() and p.get("product_type", "").lower() == proj["product_type"].lower():
                    engineer_name = p.get("owner")
                    owner_shift = p.get("owner_shift")
                    if owner_shift:
                        shift = owner_shift
                        shift_name = SHIFTS[owner_shift]["name"]
                    break
            if engineer_name:
                break

        daily = []
        for day_data in proj_coverage:
            for p in day_data.get("projects", []):
                if p["project_name"].lower() == proj["name"].lower() and p.get("product_type", "").lower() == proj["product_type"].lower():
                    shift_info = {}
                    for sn in [1, 2, 3]:
                        sh = p["shifts"].get(sn, {})
                        shift_info[sn] = {
                            "handler": sh.get("handler"),
                            "is_secondary": sh.get("is_secondary", False),
                        }
                    daily.append({
                        "day_num": day_data["day_num"],
                        "date": day_data["date"],
                        "day_abbr": day_data["day_abbr"],
                        "shifts": shift_info,
                    })
                    break

        proj_results.append({
            "name": proj["name"],
            "product_type": proj["product_type"],
            "manager": manager_name,
            "engineer": engineer_name,
            "shift": shift,
            "shift_name": shift_name,
            "daily": daily,
        })

    emp_project_data = {}
    for emp in matched_emps:
        emp_daily = {}
        for day_data in proj_coverage:
            for p in day_data.get("projects", []):
                for sn in [1, 2, 3]:
                    sh = p["shifts"].get(sn, {})
                    if sh.get("handler") == emp["name"]:
                        day_num = day_data["day_num"]
                        if day_num not in emp_daily:
                            emp_daily[day_num] = {
                                "date": day_data["date"],
                                "day_abbr": day_data["day_abbr"],
                                "projects": [],
                            }
                        emp_daily[day_num]["projects"].append({
                            "name": p["project_name"],
                            "product_type": p["product_type"],
                            "is_secondary": sh.get("is_secondary", False),
                            "owner": p["owner"],
                            "shift": sn,
                        })
        emp_project_data[emp["name"]] = emp_daily

    for er in emp_results:
        er["daily_projects"] = emp_project_data.get(er["name"], {})

    return jsonify({"employees": emp_results, "projects": proj_results})


# ── Summary ──────────────────────────────────────────────

@app.route("/summary", methods=["GET"])
@login_required
def summary():
    employees = db.get_employees_by_role("engineer")
    all_projects = db.get_all_projects()

    today = date.today()
    cy, cm = get_app_context_ym()
    year = request.args.get("year", cy, type=int)
    month = request.args.get("month", cm, type=int)

    shift_assignments = {}
    proj_coverage = []
    proj_warnings = []

    if employees:
        night_counts = db.get_night_shift_counts()
        try:
            _, _, shift_assignments = _generate_roster_with_saved_month(
                employees, year, month, night_counts
            )
        except Exception:
            shift_assignments = {}

        if all_projects and shift_assignments:
            try:
                proj_coverage, proj_warnings = generate_project_coverage(
                    all_projects, employees, shift_assignments, year, month, _get_leave_dates(year, month)
                )
            except Exception:
                pass

    month_name = calendar.month_name[month]
    month_names = [calendar.month_name[m] for m in range(1, 13)]
    year_lo = min(today.year - 1, cy - 1)
    year_hi = max(today.year + 3, cy + 2)

    return render_template("summary.html",
                           app_name=APP_NAME,
                           employees=employees,
                           all_projects=all_projects,
                           proj_coverage=proj_coverage,
                           shift_assignments=shift_assignments,
                           month_name=month_name,
                           month_names=month_names,
                           year=year,
                           month=month,
                           current_year=today.year,
                           year_range_lo=year_lo,
                           year_range_hi=year_hi)


# ── Manual Shift Editing ─────────────────────────────────

@app.route("/get_shifts", methods=["GET"])
@login_required
def get_shifts():
    """
    Return shift assignments for the Edit Shifts UI.

    Starts from the auto-generated roster for that month (same engine as Generate),
    then overlays any rows saved in rotation_history so manual edits still win.
    Engineers with no saved row show the generated shift; empty months show full roster.
    """
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        return jsonify({"error": "year and month required"}), 400
    engineers = db.get_employees_by_role("engineer")
    saved = dict(db.get_shift_assignments_for_month(year, month) or {})
    merged = {}
    if len(engineers) >= 2:
        night_counts = db.get_night_shift_counts()
        try:
            _, _, gen_assign = _generate_roster_with_saved_month(
                engineers, year, month, night_counts
            )
            merged = dict(gen_assign)
        except Exception:
            merged = {}
    merged.update(saved)
    for e in engineers:
        if is_pinned_shift_1(e["name"]):
            merged[e["name"]] = 1
    return jsonify({"assignments": merged})


@app.route("/save_shifts", methods=["GET", "POST", "OPTIONS"])
@login_required
def save_shifts():
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "GET":
        return redirect(url_for("index"))
    if not g.user or g.user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid data"}), 400
    year = data.get("year")
    month = data.get("month")
    assignments = data.get("assignments", {})
    if not year or not month:
        return jsonify({"error": "year and month required"}), 400
    # Convert shift values to int
    assignments = {k: int(v) for k, v in assignments.items() if v}
    all_emps = db.get_all_employees()
    for e in all_emps:
        if e.get("emp_role") == "engineer" and is_pinned_shift_1(e["name"]):
            assignments[e["name"]] = 1

    engineers = db.get_employees_by_role("engineer")

    if len(engineers) >= 2:
        complete = _complete_engineer_shifts_for_save(engineers, assignments, year, month)

        engineers = db.get_employees_by_role("engineer")
        night_counts = db.get_night_shift_counts()
        try:
            roster, warnings, gen_assign = generate_roster_from_manual_assignments(
                engineers,
                year,
                month,
                complete,
                prev_month_night_ids=_prev_month_night_ids(year, month),
            )
        except (ValueError, KeyError):
            # Belt-and-suspenders: accept any headcount per shift (e.g. 8 on shift 2).
            roster, warnings, gen_assign = _generate_roster(
                engineers,
                year,
                month,
                night_counts,
                predefined=complete,
                relax_fixed_caps=True,
            )

        prepared, prep_warnings = prepare_employees_for_roster_month(engineers, year, month)
        for pe in prepared:
            db.snapshot_monthly_working_pattern(pe["id"], year, month, pe["working_days"])
        if prep_warnings:
            warnings = list(warnings or []) + list(prep_warnings)

        final_rot = dict(assignments)
        for e in engineers:
            final_rot[e["name"]] = gen_assign[e["name"]]
        db.save_all_rotations(final_rot, all_emps, year, month)

        engineers = db.get_employees_by_role("engineer")
        summary = get_roster_summary(engineers, gen_assign)
        summary_json = {str(k): v for k, v in summary.items()}
        roster_data = _build_roster_data(engineers, roster, gen_assign, year, month)

        if db.get_saved_roster_data(year, month):
            all_projects = db.get_all_projects()
            proj_coverage, proj_warnings = generate_project_coverage(
                all_projects, engineers, gen_assign, year, month, _get_leave_dates(year, month)
            )
            excel_output = generate_excel(
                roster, warnings, gen_assign, engineers, year, month,
                proj_coverage, proj_warnings,
            )
            db.save_roster_excel(year, month, excel_output.read())
            db.save_roster_data(
                year,
                month,
                json.dumps(
                    {
                        "roster_data": roster_data,
                        "warnings": warnings,
                        "summary": summary_json,
                        "shift_assignments": gen_assign,
                    }
                ),
            )

        draft = session.get("draft_roster")
        if draft and draft.get("year") == year and draft.get("month") == month:
            session["draft_roster"] = {
                "year": year,
                "month": month,
                "roster_data": roster_data,
                "warnings": warnings,
                "summary": summary_json,
                "shift_assignments": gen_assign,
            }

        return jsonify({"success": True, "roster_regenerated": True})

    db.save_all_rotations(assignments, all_emps, year, month)
    return jsonify({"success": True, "roster_regenerated": False})


@app.route("/clear_shifts", methods=["GET", "POST", "OPTIONS"])
@login_required
def clear_shifts():
    if request.method == "OPTIONS":
        return "", 204
    if request.method == "GET":
        return redirect(url_for("index"))
    if not g.user or g.user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data = request.get_json() or {}
    year = data.get("year")
    month = data.get("month")
    if not year or not month:
        return jsonify({"error": "year and month required"}), 400
    db.clear_shifts_for_month(year, month)
    return jsonify({"success": True})


# ── Authentication ───────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if g.user:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        # Password login is reserved for the admin account only (emergency access).
        # All other users must sign in with Microsoft.
        if username.lower() != "admin":
            flash("Please use Microsoft login to sign in.", "info")
            return redirect(url_for("login"))
        user = db.get_user_by_username(username)
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            next_page = request.form.get("next") or url_for("index")
            return redirect(next_page)
        flash("Invalid username or password.", "danger")
    return render_template("login.html", app_name=APP_NAME,
                           next=request.args.get("next", ""),
                           azure_enabled=AZURE_ENABLED)


@app.route("/login/azure")
def login_azure():
    if not AZURE_ENABLED:
        flash("Azure AD login is not configured.", "danger")
        return redirect(url_for("login"))
    msal_app = _build_msal_app()
    redirect_uri = AZURE_REDIRECT_URI or url_for("azure_callback", _external=True)
    auth_url = msal_app.get_authorization_request_url(
        AZURE_SCOPE, redirect_uri=redirect_uri
    )
    return redirect(auth_url)


@app.route("/login/azure/callback")
def azure_callback():
    if not AZURE_ENABLED:
        return redirect(url_for("login"))

    code = request.args.get("code")
    if not code:
        flash("Azure AD login failed — no authorization code.", "danger")
        return redirect(url_for("login"))

    msal_app = _build_msal_app()
    redirect_uri = AZURE_REDIRECT_URI or url_for("azure_callback", _external=True)
    result = msal_app.acquire_token_by_authorization_code(
        code, scopes=AZURE_SCOPE, redirect_uri=redirect_uri
    )

    if "error" in result:
        flash(f"Azure AD login failed: {result.get('error_description', result['error'])}", "danger")
        return redirect(url_for("login"))

    azure_user = result.get("id_token_claims", {})
    email = azure_user.get("preferred_username", "").strip().lower()
    display_name = azure_user.get("name", "").strip()

    if not email:
        flash("Could not retrieve email from Microsoft. Please try again.", "danger")
        return redirect(url_for("login"))

    # Step 1: Already migrated — email is the username
    user = db.get_user_by_username(email)

    # Step 2: First-ever Microsoft login — find legacy short username and migrate
    if not user:
        legacy_username = _LEGACY_USERNAME_MAP.get(email)
        if legacy_username:
            user = db.get_user_by_username_ci(legacy_username)
            if user:
                db.update_user_username(user["id"], email)
                user = db.get_user_by_id(user["id"])
                app.logger.info("Migrated username '%s' → '%s'", legacy_username, email)

    # Step 3: Completely new user — create account with user role
    if not user:
        matched_emp = db.auto_link_user(display_name)
        emp_id = matched_emp["id"] if matched_emp else None
        uid = db.add_user(
            username=email,
            password_hash=generate_password_hash(os.urandom(32).hex()),
            full_name=display_name,
            role="user",
            employee_id=emp_id,
        )
        user = db.get_user_by_id(uid) if uid else db.get_user_by_username(email)

    if user:
        session.clear()
        session["user_id"] = user["id"]
        flash(f"Welcome, {display_name or email}!", "success")
        return redirect(url_for("index"))

    flash("Login failed. Please contact your administrator.", "danger")
    return redirect(url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if g.user:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        error = None
        if not username or len(username) < 3:
            error = "Username must be at least 3 characters."
        elif not password or len(password) < 6:
            error = "Password must be at least 6 characters."
        elif password != confirm:
            error = "Passwords do not match."
        elif db.get_user_by_username(username):
            error = f"Username '{username}' is already taken."
        if error:
            flash(error, "danger")
        else:
            matched_emp = db.auto_link_user(full_name)
            emp_id = matched_emp["id"] if matched_emp else None
            db.add_user(username, generate_password_hash(password), full_name, role="user", employee_id=emp_id)
            if matched_emp:
                role_label = ROLE_LABELS.get(matched_emp.get("emp_role", "engineer"), "Employee")
                flash(f"Account created! Linked to {role_label}: {matched_emp['name']}.", "success")
            else:
                flash("Account created! No matching employee found -- view-only access.", "info")
            return redirect(url_for("login"))
    return render_template("signup.html", app_name=APP_NAME)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = g.user
    if request.method == "POST":
        action = request.form.get("action")
        if action == "update_profile":
            full_name = request.form.get("full_name", "").strip()
            db.update_user_profile(user["id"], full_name)
            flash("Profile updated.", "success")
            return redirect(url_for("profile"))
        elif action == "change_password":
            current_pw = request.form.get("current_password", "")
            new_pw = request.form.get("new_password", "")
            confirm_pw = request.form.get("confirm_password", "")
            if not check_password_hash(user["password_hash"], current_pw):
                flash("Current password is incorrect.", "danger")
            elif len(new_pw) < 6:
                flash("New password must be at least 6 characters.", "danger")
            elif new_pw != confirm_pw:
                flash("Passwords do not match.", "danger")
            else:
                db.update_user_password(user["id"], generate_password_hash(new_pw))
                flash("Password changed successfully.", "success")
            return redirect(url_for("profile"))
    # Refresh user from DB to show latest data
    user = db.get_user_by_id(user["id"])
    return render_template("profile.html", app_name=APP_NAME, user=user)


# ── Leave Tracker ────────────────────────────────────────

@app.route("/leaves")
@login_required
def leaves():
    today = date.today()
    cy, cm = get_app_context_ym()
    year = request.args.get("year", cy, type=int)
    month = request.args.get("month", cm, type=int)
    month_names = [calendar.month_name[m] for m in range(1, 13)]

    is_admin = g.user and g.user.get("role") == "admin"
    linked_emp = db.get_linked_employee(g.user["id"]) if g.user else None
    linked_emp_id = linked_emp["id"] if linked_emp else None

    month_leaves = db.get_leaves_for_month(year, month)
    pending_requests = db.get_pending_requests() if is_admin else []
    engineers = db.get_employees_by_role("engineer")

    if not is_admin and linked_emp_id:
        month_leaves = [l for l in month_leaves if l["employee_id"] == linked_emp_id]

    balances = db.get_all_balances_for_year(year)
    bal_map = {b["employee_id"]: b for b in balances}
    for emp in engineers:
        if emp["id"] not in bal_map:
            db.get_or_create_balance(emp["id"], year)
    balances = db.get_all_balances_for_year(year)

    if not is_admin and linked_emp_id:
        balances = [b for b in balances if b["employee_id"] == linked_emp_id]

    shift_assignments = {}
    if engineers:
        night_counts = db.get_night_shift_counts()
        try:
            _, _, shift_assignments = _generate_roster_with_saved_month(
                engineers, year, month, night_counts
            )
        except Exception:
            shift_assignments = {}

    strength = db.get_shift_strength(year, month, shift_assignments) if is_admin else {}

    year_lo = min(today.year - 1, cy - 1)
    year_hi = max(today.year + 3, cy + 2)

    return render_template("leaves.html",
                           app_name=APP_NAME,
                           month_leaves=month_leaves,
                           pending_requests=pending_requests,
                           engineers=engineers,
                           balances=balances,
                           strength=strength,
                           shift_assignments=shift_assignments,
                           shifts=SHIFTS,
                           month_names=month_names,
                           year=year,
                           month=month,
                           current_year=today.year,
                           year_range_lo=year_lo,
                           year_range_hi=year_hi,
                           is_admin=is_admin,
                           linked_employee=linked_emp)


@app.route("/add_leave", methods=["POST"])
@admin_required
def add_leave_route():
    emp_id = request.form.get("employee_id", type=int)
    leave_date = request.form.get("leave_date")
    leave_type = request.form.get("leave_type", "planned")
    reason = request.form.get("reason", "")

    if not emp_id or not leave_date:
        flash("Employee and date are required.", "danger")
        return redirect(url_for("leaves"))

    lid = db.add_leave(emp_id, leave_date, leave_type, reason, approved_by=g.user["username"])
    if lid is None:
        flash("Leave already exists for that date.", "warning")
    else:
        emp = db.get_employee_by_id(emp_id)
        yr = int(leave_date.split("-")[0])
        db.increment_leave_used(emp_id, yr, leave_type)
        flash(f"Leave added for {emp['name'] if emp else 'employee'} on {leave_date}.", "success")

    return redirect(url_for("leaves"))


@app.route("/cancel_leave/<int:leave_id>", methods=["POST"])
@admin_required
def cancel_leave_route(leave_id):
    row = db.cancel_leave(leave_id)
    if row:
        yr = row["leave_date"].year if hasattr(row["leave_date"], "year") else int(str(row["leave_date"])[:4])
        db.decrement_leave_used(row["employee_id"], yr, row["leave_type"])
        flash("Leave cancelled.", "success")
    return redirect(url_for("leaves"))


@app.route("/approve_leave/<int:request_id>", methods=["POST"])
@admin_required
def approve_leave_route(request_id):
    req = db.get_leave_request_by_id(request_id)
    if not req:
        flash("Request not found.", "danger")
        return redirect(url_for("leaves"))

    result = db.approve_leave_request(request_id, g.user["username"])
    if result:
        emp_id, leave_date, leave_type, reason = result
        lid = db.add_leave(emp_id, leave_date, leave_type, reason, approved_by=g.user["username"])
        if lid:
            yr = leave_date.year if hasattr(leave_date, "year") else int(str(leave_date)[:4])
            db.increment_leave_used(emp_id, yr, leave_type)
        flash(f"Leave approved for {req['employee_name']}.", "success")

    return redirect(url_for("leaves"))


@app.route("/reject_leave/<int:request_id>", methods=["POST"])
@admin_required
def reject_leave_route(request_id):
    db.reject_leave_request(request_id, g.user["username"])
    flash("Leave request rejected.", "success")
    return redirect(url_for("leaves"))


@app.route("/request_leave", methods=["POST"])
@login_required
def request_leave():
    emp_id = request.form.get("employee_id", type=int)
    leave_date = request.form.get("leave_date")
    leave_type = request.form.get("leave_type", "planned")
    reason = request.form.get("reason", "")

    is_admin = g.user and g.user.get("role") == "admin"
    if not is_admin and g.user.get("employee_id"):
        emp_id = g.user["employee_id"]

    if not emp_id or not leave_date:
        flash("Please select employee and date.", "danger")
        return redirect(url_for("leaves"))

    from datetime import datetime
    req_date = datetime.strptime(leave_date, "%Y-%m-%d").date()
    days_ahead = (req_date - date.today()).days

    if days_ahead < 2 and leave_type == "planned":
        flash("Planned leave must be requested at least 2 days in advance.", "warning")
        return redirect(url_for("leaves"))

    db.add_leave_request(emp_id, g.user["id"], leave_date, leave_type, reason)
    flash(f"Leave request submitted for {leave_date}.", "success")
    return redirect(url_for("leaves"))


@app.route("/leave_impact", methods=["GET"])
@login_required
def leave_impact():
    emp_id = request.args.get("employee_id", type=int)
    leave_date = request.args.get("date")

    if not emp_id or not leave_date:
        return jsonify({"error": "employee_id and date required"}), 400

    emp = db.get_employee_by_id(emp_id)
    if not emp:
        return jsonify({"error": "Employee not found"}), 404

    from datetime import datetime
    d = datetime.strptime(leave_date, "%Y-%m-%d").date()
    year, month = d.year, d.month

    engineers = db.get_employees_by_role("engineer")
    night_counts = db.get_night_shift_counts()
    try:
        _, _, shift_assignments = _generate_roster_with_saved_month(
            engineers, year, month, night_counts
        )
    except Exception:
        shift_assignments = {}

    emp_shift = shift_assignments.get(emp["name"])

    leave_map = _get_leave_dates(year, month)
    leave_map.setdefault(emp["name"], []).append(leave_date)

    strength = db.get_shift_strength(year, month, shift_assignments)
    day_strength = strength.get(d.day, {})
    if emp_shift and day_strength.get(emp_shift, 0) > 0:
        day_strength[emp_shift] -= 1

    all_projects = db.get_all_projects()
    emp_projects = [p for p in all_projects if p["employee_name"] == emp["name"]]

    balance = db.get_or_create_balance(emp_id, year)
    used = (balance.get("planned_used", 0) + balance.get("sick_used", 0) + balance.get("emergency_used", 0))
    remaining = balance.get("total_allowed", 24) - used

    return jsonify({
        "employee": emp["name"],
        "shift": emp_shift,
        "shift_strength": day_strength,
        "projects_affected": len(emp_projects),
        "project_names": [p["name"] for p in emp_projects],
        "leaves_remaining": remaining,
        "balance": {
            "total": balance.get("total_allowed", 24),
            "planned": balance.get("planned_used", 0),
            "sick": balance.get("sick_used", 0),
            "emergency": balance.get("emergency_used", 0),
        }
    })


# ── Agent Bot ────────────────────────────────────────────

import re as _re

_MONTH_MAP = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}


def _bot_parse_date(query):
    today = date.today()
    for mname, mnum in sorted(_MONTH_MAP.items(), key=lambda x: -len(x[0])):
        if mname in query:
            m = _re.search(r'(\d{1,2})\s*(?:st|nd|rd|th)?\s*(?:of\s*)?' + mname, query)
            if not m:
                m = _re.search(mname + r'\s*(\d{1,2})', query)
            if m:
                d = int(m.group(1))
                y = today.year
                yr = _re.search(r'20\d{2}', query)
                if yr:
                    y = int(yr.group())
                elif mnum < today.month:
                    y = today.year + 1
                try:
                    date(y, mnum, d)
                    return (y, mnum, d)
                except ValueError:
                    pass
            return None
    return None


def _bot_parse_shift(query):
    for pat in [r'shift\s*(\d)', r'(\d)\s*(?:st|nd|rd|th)\s*shift', r'\bs(\d)\b']:
        m = _re.search(pat, query)
        if m:
            s = int(m.group(1))
            if 1 <= s <= 3:
                return s
    return None


def _bot_get_coverage(year, month, engineers, shift_assignments):
    all_proj = db.get_all_projects()
    if engineers and shift_assignments and all_proj:
        cov, _ = generate_project_coverage(all_proj, engineers, shift_assignments, year, month, _get_leave_dates(year, month))
        return cov
    return []


@app.route("/bot", methods=["POST"])
@login_required
def bot():
    data = request.get_json()
    q = (data.get("query") or "").strip().lower()
    if not q:
        return jsonify({"answer": "Please ask a question."})

    today = date.today()

    pd = _bot_parse_date(q)
    asked_shift = _bot_parse_shift(q)
    q_year, q_month, q_day = (pd[0], pd[1], pd[2]) if pd else (today.year, today.month, None)

    engineers = db.get_employees_by_role("engineer")
    all_projects = db.get_all_projects()
    all_people = db.get_all_employees()
    shift_leads = db.get_employees_by_role("shift_lead")
    managers = db.get_employees_by_role("manager")

    shift_assignments = {}
    if engineers:
        sr = db.get_saved_roster_data(q_year, q_month)
        if sr:
            shift_assignments = json.loads(sr["roster_json"]).get("shift_assignments", {})
        else:
            nc = db.get_night_shift_counts()
            _, _, shift_assignments = _generate_roster_with_saved_month(
                engineers, q_year, q_month, nc
            )

    emp_match = None
    for e in sorted(all_people, key=lambda x: -len(x["name"])):
        if e["name"].lower() in q:
            emp_match = e
            break

    proj_match = None
    for p in sorted(all_projects, key=lambda x: -len(x["name"])):
        if p["name"].lower() in q:
            proj_match = p
            break

    # Team counts
    if any(w in q for w in ["how many engineer", "total engineer"]):
        return jsonify({"answer": f"There are {len(engineers)} migration engineers."})
    if any(w in q for w in ["how many lead", "total lead"]):
        return jsonify({"answer": f"There are {len(shift_leads)} shift leads."})
    if any(w in q for w in ["how many manager", "total manager"]):
        return jsonify({"answer": f"There are {len(managers)} migration managers."})
    if any(w in q for w in ["how many project", "total project"]):
        return jsonify({"answer": f"There are {len(set(p['name'] for p in all_projects))} projects."})

    # List queries
    if any(w in q for w in ["list engineer", "all engineer", "show engineer"]):
        return jsonify({"answer": "Engineers: " + ", ".join(e["name"] for e in engineers)})
    if any(w in q for w in ["list lead", "all lead", "show lead"]):
        return jsonify({"answer": "Shift Leads: " + (", ".join(e["name"] for e in shift_leads) or "None")})
    if any(w in q for w in ["list manager", "all manager", "show manager"]):
        return jsonify({"answer": "Managers: " + (", ".join(e["name"] for e in managers) or "None")})
    if any(w in q for w in ["list project", "all project", "show project"]):
        return jsonify({"answer": "Projects: " + ", ".join(sorted(set(p["name"] for p in all_projects)))})

    # Project + date/shift queries (highest priority)
    if proj_match:
        pname = proj_match["name"]
        mgr_name = None
        eng_name = None
        eng_shift = None
        for p in all_projects:
            if p["name"] == pname:
                mgr = next((m for m in managers if m["id"] == p["employee_id"]), None)
                if mgr:
                    mgr_name = mgr["name"]
                emp = next((e for e in engineers if e["id"] == p["employee_id"]), None)
                if emp:
                    eng_name = emp["name"]
                    eng_shift = shift_assignments.get(emp["name"])

        coverage = _bot_get_coverage(q_year, q_month, engineers, shift_assignments)

        if q_day:
            mn = calendar.month_name[q_month]
            for day_data in coverage:
                if day_data.get("day_num") == q_day:
                    for p in day_data.get("projects", []):
                        if p["project_name"].lower() == pname.lower():
                            shifts = p.get("shifts", {})
                            if asked_shift:
                                sh = shifts.get(asked_shift) or shifts.get(str(asked_shift)) or {}
                                h = sh.get("handler") or "No one available"
                                return jsonify({"answer": f"{pname} on {mn} {q_day} ({day_data['day_abbr']}), {SHIFTS[asked_shift]['name']}: {h}" + (f" | Manager: {mgr_name}" if mgr_name else "")})
                            else:
                                parts = []
                                for sn in [1, 2, 3]:
                                    sh = shifts.get(sn) or shifts.get(str(sn)) or {}
                                    h = sh.get("handler") or "—"
                                    parts.append(f"S{sn}: {h}")
                                return jsonify({"answer": f"{pname} on {mn} {q_day} ({day_data['day_abbr']}): {' | '.join(parts)}" + (f" | Manager: {mgr_name}" if mgr_name else "")})
                    return jsonify({"answer": f"No data for {pname} on {mn} {q_day}."})

        if asked_shift and coverage:
            mn = calendar.month_name[q_month]
            for day_data in coverage:
                for p in day_data.get("projects", []):
                    if p["project_name"].lower() == pname.lower():
                        shifts = p.get("shifts", {})
                        sh = shifts.get(asked_shift) or shifts.get(str(asked_shift)) or {}
                        h = sh.get("handler")
                        if h:
                            return jsonify({"answer": f"{pname}, {SHIFTS[asked_shift]['name']} ({mn} {q_year}): {h}" + (f" | Manager: {mgr_name}" if mgr_name else "")})
                break
            return jsonify({"answer": f"No handler found for {pname} in {SHIFTS[asked_shift]['name']}."})

        if any(w in q for w in ["who", "handle", "working", "manage", "assign"]) and coverage:
            mn = calendar.month_name[q_month]
            for day_data in coverage:
                for p in day_data.get("projects", []):
                    if p["project_name"].lower() == pname.lower():
                        shifts = p.get("shifts", {})
                        parts = []
                        for sn in [1, 2, 3]:
                            sh = shifts.get(sn) or shifts.get(str(sn)) or {}
                            h = sh.get("handler")
                            if h:
                                parts.append(f"{SHIFTS[sn]['name']}: {h}")
                        answer = f"{pname} ({mn} {q_year}): {' | '.join(parts)}" if parts else f"No coverage data for {pname}"
                        if mgr_name:
                            answer += f" | Manager: {mgr_name}"
                        return jsonify({"answer": answer})
                break

        answer = f"{pname} (Type: {proj_match['product_type']})"
        if eng_name:
            answer += f" | Engineer: {eng_name}"
            if eng_shift:
                answer += f" ({SHIFTS[eng_shift]['name']})"
        if mgr_name:
            answer += f" | Manager: {mgr_name}"
        return jsonify({"answer": answer})

    # Shift roster queries (no project)
    if asked_shift and not emp_match and any(w in q for w in ["who", "working", "shift"]):
        emps = [n for n, s in shift_assignments.items() if s == asked_shift]
        mn = calendar.month_name[q_month]
        if q_day:
            d = date(q_year, q_month, q_day)
            dn = DAY_NAMES[d.weekday()]
            lm = _get_leave_dates(q_year, q_month)
            el = {e["name"]: e for e in engineers}
            working = [n for n in emps if el.get(n) and dn in el[n]["working_days"] and d.strftime("%Y-%m-%d") not in lm.get(n, [])]
            return jsonify({"answer": f"{SHIFTS[asked_shift]['name']} on {mn} {q_day} ({d.strftime('%a')}): {', '.join(working) if working else 'No one'}"})
        return jsonify({"answer": f"{SHIFTS[asked_shift]['name']} for {mn} {q_year}: {', '.join(emps) if emps else 'None'}"})

    # Employee queries
    if emp_match:
        name = emp_match["name"]
        role = emp_match.get("emp_role", "engineer")
        rl = ROLE_LABELS.get(role, "Employee")
        shift = shift_assignments.get(name)
        sn = SHIFTS[shift]["name"] if shift else "Not assigned"
        ep = [p["name"] for p in all_projects if p["employee_id"] == emp_match["id"]]

        if q_day:
            d = date(q_year, q_month, q_day)
            dn = DAY_NAMES[d.weekday()]
            ds = d.strftime("%Y-%m-%d")
            lm = _get_leave_dates(q_year, q_month)
            on_leave = ds in lm.get(name, [])
            is_off = dn not in emp_match.get("working_days", [])
            status = "on leave" if on_leave else ("off (weekly)" if is_off else "working")
            mn = calendar.month_name[q_month]

            coverage = _bot_get_coverage(q_year, q_month, engineers, shift_assignments)
            handling = []
            for dd in coverage:
                if dd.get("day_num") == q_day:
                    for p in dd.get("projects", []):
                        shifts_data = p.get("shifts", {})
                        for s in [1, 2, 3]:
                            sh = shifts_data.get(s) or shifts_data.get(str(s)) or {}
                            if sh.get("handler") == name:
                                handling.append(p["project_name"])
                    break

            answer = f"{name} on {mn} {q_day} ({d.strftime('%a')}): {status}, {sn}"
            if handling:
                answer += f" | Projects: {', '.join(set(handling))}"
            return jsonify({"answer": answer})

        if any(w in q for w in ["shift", "which shift"]):
            return jsonify({"answer": f"{name} is in {sn} for {calendar.month_name[q_month]} {q_year}."})
        if any(w in q for w in ["project", "what project"]):
            return jsonify({"answer": f"{name} handles: {', '.join(ep)}" if ep else f"{name} has no projects."})
        if any(w in q for w in ["leave", "balance"]):
            b = db.get_or_create_balance(emp_match["id"], q_year)
            used = b["planned_used"] + b["sick_used"] + b["emergency_used"]
            rem = b["total_allowed"] - used
            return jsonify({"answer": f"{name}: {rem}/{b['total_allowed']} leaves remaining (P:{b['planned_used']}, S:{b['sick_used']}, E:{b['emergency_used']})"})
        if any(w in q for w in ["off day", "week off", "working day"]):
            days = emp_match.get("working_days", [])
            off = [d for d in DAY_NAMES if d not in days]
            return jsonify({"answer": f"{name} works {', '.join(d[:3] for d in days)}. Off: {', '.join(d[:3] for d in off)}"})

        info = f"{name} ({rl}), {sn}"
        days = emp_match.get("working_days", [])
        off = [d for d in DAY_NAMES if d not in days]
        info += f" | Works: {', '.join(d[:3] for d in days)} | Off: {', '.join(d[:3] for d in off)}"
        if ep:
            info += f" | Projects: {', '.join(ep)}"
        return jsonify({"answer": info})

    if any(w in q for w in ["help", "what can", "how to", "hi", "hello"]):
        return jsonify({"answer": "Ask me about:\n• 'Who handles Washington Post on 6th April?'\n• 'Washington Post shift 2 on 6th april'\n• 'What shift is Arun in?'\n• 'Arun on 10th april'\n• 'Who is in shift 2 on april 10?'\n• 'How many leaves does David have?'\n• 'List all projects'"})

    return jsonify({"answer": "I couldn't find a match. Try mentioning an employee name, project name, or date. Type 'help' for examples."})


# ── Delta Calendar ───────────────────────────────────────

@app.route("/delta")
@login_required
def delta():
    cy, cm = get_app_context_ym()
    all_projects = db.get_all_projects()
    all_people = db.get_all_employees()
    emp_role_map = {e["name"]: e.get("emp_role", "engineer") for e in all_people}

    excluded = db.get_excluded_projects()

    seen = set()
    unique_projects = []
    for p in all_projects:
        if emp_role_map.get(p["employee_name"]) != "engineer":
            continue
        key = (p["name"], p["product_type"])
        if key in excluded:
            continue
        if key not in seen:
            seen.add(key)
            unique_projects.append({"name": p["name"], "product_type": p["product_type"]})
    unique_projects.sort(key=lambda p: (p["name"].lower(), p["product_type"].lower()))

    proj_manager_map = {}
    for p in all_projects:
        if emp_role_map.get(p["employee_name"]) == "manager":
            proj_manager_map[p["name"]] = p["employee_name"]

    managers = [e["name"] for e in all_people if e.get("emp_role") == "manager"]
    managers.sort()

    delta_events = db.get_all_delta_events()
    for ev in delta_events:
        ev["assignments"] = db.get_delta_assignments(ev["id"])

    all_types = sorted({p["product_type"] for p in all_projects}) if all_projects else []

    final_validation_projects = db.get_project_lifecycle_by_status("final_validation")
    decommissioned_projects = db.get_project_lifecycle_by_status("decommissioned")

    # Build lookup keys so the template can override badge colour from project_lifecycle
    # (authoritative source for phase) rather than delta_events.delta_status alone
    final_validation_keys = [
        f"{p['project_name']}||{p['product_type']}" for p in final_validation_projects
    ]
    decommissioned_keys = [
        f"{p['project_name']}||{p['product_type']}" for p in decommissioned_projects
    ]

    today = date.today()
    today_str = today.isoformat()
    from datetime import timedelta
    week_start = today - timedelta(days=today.weekday())   # Monday
    week_end   = week_start + timedelta(days=6)             # Sunday

    # Active deltas (end_date >= today) appear exactly once in Running with
    # all their assignments.  Completed deltas (end_date < today) are bucketed
    # into history by the week their assignments fall in.  This prevents a
    # multi-week active delta from appearing in both Running and History.
    current_week_events = []
    history_by_week = {}  # {week_monday_date: [event_copies_with_filtered_assignments]}
    for ev in delta_events:
        ev_end_date = date.fromisoformat(str(ev.get("end_date", today_str)))

        if ev_end_date >= today:
            # Still active — show once in Running with all assignments
            current_week_events.append(ev)
        else:
            # Completed — bucket each week's assignments into history
            week_assignments = {}
            for a in ev.get("assignments", []):
                a_date = date.fromisoformat(str(a["assignment_date"]))
                a_week_mon = a_date - timedelta(days=a_date.weekday())
                week_assignments.setdefault(a_week_mon, []).append(a)
            for wk_mon, wk_assigns in sorted(week_assignments.items()):
                history_by_week.setdefault(wk_mon, []).append(
                    {**ev, "assignments": wk_assigns}
                )

    sorted_history = [
        (ws.isoformat(), (ws + timedelta(days=6)).isoformat(), evts)
        for ws, evts in sorted(history_by_week.items(), key=lambda x: x[0], reverse=True)
    ]

    return render_template(
        "delta.html",
        app_name=APP_NAME,
        unique_projects=unique_projects,
        proj_manager_map=proj_manager_map,
        managers=managers,
        delta_events=delta_events,
        product_types=all_types,
        year=cy,
        month=cm,
        final_validation_projects=final_validation_projects,
        decommissioned_projects=decommissioned_projects,
        final_validation_keys=final_validation_keys,
        decommissioned_keys=decommissioned_keys,
        today_str=today_str,
        week_start=week_start.isoformat(),
        week_end=week_end.isoformat(),
        current_week_events=current_week_events,
        sorted_history=sorted_history,
    )


@app.route("/delta/preview", methods=["POST"])
@login_required
def delta_preview():
    data = request.get_json()
    project_name = (data.get("project_name") or "").strip()
    product_type = (data.get("product_type") or "").strip()
    start_str = (data.get("start_date") or "").strip()
    end_str = (data.get("end_date") or "").strip()

    if not project_name or not product_type or not start_str or not end_str:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        start_d = date.fromisoformat(start_str)
        end_d = date.fromisoformat(end_str)
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    if end_d < start_d:
        return jsonify({"error": "End date must be on or after start date"}), 400

    from datetime import timedelta
    all_dates = []
    cur_d = start_d
    while cur_d <= end_d:
        all_dates.append(cur_d)
        cur_d += timedelta(days=1)

    from collections import defaultdict
    dates_by_month = defaultdict(list)
    for d in all_dates:
        dates_by_month[(d.year, d.month)].append(d)

    all_projects = db.get_all_projects()
    _excl = db.get_excluded_projects()
    all_projects = [p for p in all_projects if (p["name"], p["product_type"]) not in _excl]
    engineers = db.get_employees_by_role("engineer")

    result_days = []
    for (yr, mo), month_dates in sorted(dates_by_month.items()):
        try:
            _, _, shift_assignments = _generate_roster_with_saved_month(engineers, yr, mo)
        except Exception:
            shift_assignments = {}

        leave_dates = _get_leave_dates(yr, mo)

        try:
            coverage, _ = generate_project_coverage(
                all_projects, engineers, shift_assignments, yr, mo, leave_dates
            )
        except Exception:
            coverage = []

        cov_by_day = {c["day_num"]: c for c in coverage}

        prepared_emps, _ = prepare_employees_for_roster_month(engineers, yr, mo)
        emp_lookup = {e["name"]: e for e in prepared_emps}

        for d in month_dates:
            day_cov = cov_by_day.get(d.day, {})
            proj_entry = None
            for p in day_cov.get("projects", []):
                if p["project_name"] == project_name and p["product_type"] == product_type:
                    proj_entry = p
                    break

            date_str = d.strftime("%Y-%m-%d")
            shifts_out = {}
            for sn in [1, 2, 3]:
                auto_handler = None
                if proj_entry:
                    sh = proj_entry.get("shifts", {}).get(sn, {})
                    auto_handler = sh.get("handler")

                eligible = []
                for emp in prepared_emps:
                    if shift_assignments.get(emp["name"]) != sn:
                        continue
                    if product_type not in emp.get("content_types", []):
                        continue
                    if not is_emp_scheduled_work_day(emp, d):
                        continue
                    if date_str in leave_dates.get(emp["name"], []):
                        continue
                    eligible.append(emp["name"])
                eligible.sort()

                if auto_handler and auto_handler not in eligible:
                    eligible.insert(0, auto_handler)

                shifts_out[str(sn)] = {"auto": auto_handler, "eligible": eligible}

            result_days.append({
                "date_str": date_str,
                "day_name": d.strftime("%a, %b %d"),
                "shifts": shifts_out,
            })

    return jsonify({"days": result_days})


@app.route("/delta/save", methods=["POST"])
@login_required
def delta_save():
    project_name = (request.form.get("project_name") or "").strip()
    product_type = (request.form.get("product_type") or "").strip()
    start_date = (request.form.get("start_date") or "").strip()
    end_date = (request.form.get("end_date") or "").strip()

    if not project_name or not product_type or not start_date or not end_date:
        flash("Missing required fields.", "danger")
        return redirect(url_for("delta"))

    manager_name = (request.form.get("manager_name") or "").strip()
    start_time = (request.form.get("start_time") or "").strip()
    end_time = (request.form.get("end_time") or "").strip()
    created_by = g.user.get("username", "") if g.user else ""
    delta_id = db.add_delta_event(project_name, product_type, start_date, end_date, created_by, manager_name, start_time, end_time)

    assignments = []
    for key, value in request.form.items():
        if key.startswith("assign["):
            try:
                inner = key[len("assign["):-1]
                date_part, shift_part = inner.rsplit("][", 1)
                shift_num = int(shift_part)
                engineer_name = (value or "").strip()
                if engineer_name:
                    assignments.append({
                        "date": date_part,
                        "shift_num": shift_num,
                        "engineer_name": engineer_name,
                    })
            except (ValueError, IndexError):
                continue

    if assignments:
        db.save_delta_assignments(delta_id, assignments)

    flash(f"Delta for '{project_name}' saved successfully.", "success")
    return redirect(url_for("delta"))


@app.route("/delta/download")
@login_required
def delta_download():
    from datetime import timedelta
    product_type = (request.args.get("product_type") or "").strip()
    if not product_type:
        flash("No product type specified.", "danger")
        return redirect(url_for("delta"))

    today = date.today()

    # Priority: explicit from_date/to_date  >  week (Monday)  >  current week
    from_param = (request.args.get("from_date") or "").strip()
    to_param   = (request.args.get("to_date")   or "").strip()
    week_param = (request.args.get("week")       or "").strip()

    try:
        if from_param and to_param:
            range_start = date.fromisoformat(from_param)
            range_end   = date.fromisoformat(to_param)
            if range_end < range_start:
                range_end = range_start
            label = f"{range_start.strftime('%Y%m%d')}_{range_end.strftime('%Y%m%d')}"
        elif week_param:
            range_start = date.fromisoformat(week_param)
            range_end   = range_start + timedelta(days=6)
            label = range_start.strftime('%Y%m%d')
        else:
            range_start = today - timedelta(days=today.weekday())
            range_end   = range_start + timedelta(days=6)
            label = range_start.strftime('%Y%m%d')
    except ValueError:
        flash("Invalid date parameter.", "danger")
        return redirect(url_for("delta"))

    all_events = db.get_delta_events_by_product_type(product_type)

    assignments_by_id = {}
    for ev in all_events:
        raw = db.get_delta_assignments(ev["id"])
        filtered = [
            a for a in raw
            if range_start <= date.fromisoformat(str(a["assignment_date"])) <= range_end
        ]
        if filtered:
            assignments_by_id[ev["id"]] = filtered

    events = [ev for ev in all_events if ev["id"] in assignments_by_id]

    if not events:
        flash(f"No delta assignments for '{product_type}' between {range_start} and {range_end}.", "warning")
        return redirect(url_for("delta"))

    output = generate_delta_image(events, assignments_by_id, product_type)
    filename = f"Delta_{product_type}_{label}.png"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="image/png",
    )


@app.route("/delta/delete/<int:delta_id>", methods=["POST"])
@login_required
def delta_delete(delta_id):
    db.delete_delta_event(delta_id)
    flash("Delta event deleted.", "success")
    return redirect(url_for("delta"))


@app.route("/delta/range-data")
@login_required
def delta_range_data():
    from datetime import timedelta
    from_param = (request.args.get("from_date") or "").strip()
    to_param   = (request.args.get("to_date")   or "").strip()
    pt_param   = (request.args.get("product_type") or "").strip()

    if not from_param or not to_param:
        return jsonify({"error": "Both from_date and to_date are required."}), 400

    try:
        range_start = date.fromisoformat(from_param)
        range_end   = date.fromisoformat(to_param)
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

    if range_end < range_start:
        return jsonify({"error": "End date must be on or after start date."}), 400

    if pt_param:
        all_events = db.get_delta_events_by_product_type(pt_param)
    else:
        all_events = db.get_all_delta_events()

    results = []
    for ev in all_events:
        raw = db.get_delta_assignments(ev["id"])
        filtered = sorted(
            [a for a in raw
             if range_start <= date.fromisoformat(str(a["assignment_date"])) <= range_end],
            key=lambda a: (str(a["assignment_date"]), a["shift_num"])
        )
        if filtered:
            results.append({
                "project_name": ev["project_name"],
                "product_type": ev["product_type"],
                "manager_name": ev.get("manager_name") or "",
                "start_date":   str(ev["start_date"]),
                "end_date":     str(ev["end_date"]),
                "delta_status": ev.get("delta_status") or "active",
                "assignments":  [
                    {"date": str(a["assignment_date"]),
                     "shift": a["shift_num"],
                     "engineer": a["engineer_name"]}
                    for a in filtered
                ],
            })

    return jsonify({"events": results, "from_date": from_param, "to_date": to_param, "total": len(results)})


@app.route("/delta/set-status/<int:delta_id>", methods=["POST"])
@login_required
def delta_set_status(delta_id):
    status = request.form.get("delta_status", "").strip()
    valid = {"pre_delta", "delta_issues", "final_validation"}
    if status not in valid:
        flash("Invalid status.", "danger")
        return redirect(url_for("delta"))

    db.update_delta_status(delta_id, status)

    if status == "final_validation":
        project_name = request.form.get("project_name", "").strip()
        product_type = request.form.get("product_type", "").strip()
        if project_name and product_type:
            db.upsert_project_lifecycle(
                project_name, product_type,
                migration_status="final_validation",
                final_validation_date=date.today(),
            )
            flash(f"'{project_name}' marked as Final Validation — removed from active roster.", "warning")
        else:
            flash("Status updated to Final Validation.", "warning")
    else:
        label = {"pre_delta": "Pre-delta", "delta_issues": "Delta Issues"}[status]
        flash(f"Delta status updated to {label}.", "success")

    return redirect(url_for("delta"))


@app.route("/delta/decommission", methods=["POST"])
@login_required
def delta_decommission():
    project_name = request.form.get("project_name", "").strip()
    product_type = request.form.get("product_type", "").strip()
    if not project_name or not product_type:
        flash("Missing project information.", "danger")
        return redirect(url_for("delta"))

    db.upsert_project_lifecycle(
        project_name, product_type,
        migration_status="decommissioned",
        decommission_date=date.today(),
    )
    db.update_delta_events_status_by_project(project_name, product_type, "decommissioned")
    flash(f"'{project_name}' has been decommissioned as of {date.today()}.", "success")
    return redirect(url_for("delta"))


@app.route("/delta/reactivate", methods=["POST"])
@login_required
def delta_reactivate():
    project_name = request.form.get("project_name", "").strip()
    product_type = request.form.get("product_type", "").strip()
    if not project_name or not product_type:
        flash("Missing project information.", "danger")
        return redirect(url_for("delta"))
    db.delete_project_lifecycle(project_name, product_type)
    flash(f"'{project_name}' has been reactivated and is available for new delta events.", "success")
    return redirect(url_for("delta"))


@app.route("/favicon.ico")
def favicon():
    return "", 204


# ── Drive Change Dashboard ───────────────────────────────────────────────────

@app.route("/drive-change", methods=["GET"])
@login_required
def drive_change_dashboard():
    return render_template("drive_change.html", app_name=APP_NAME)


@app.route("/drive-change/refresh", methods=["POST"])
@login_required
def drive_change_refresh():
    from drive_change_reader import fetch_drive_change_emails
    from drive_change_parser import parse_all, WINDOWS
    from drive_change_grouper import group_alerts
    from datetime import datetime, timedelta, timezone as tz

    IST = tz(timedelta(hours=5, minutes=30))
    body = request.get_json() or {}
    date_str = (body.get("date") or "").strip()
    if not date_str:
        date_str = datetime.now(IST).strftime("%Y-%m-%d")

    try:
        emails = fetch_drive_change_emails(date_str)
    except EnvironmentError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Failed to fetch emails: {e}"}), 500

    parsed = parse_all(emails)
    grouped = group_alerts(parsed)

    window_totals = {
        str(w): grouped.get(w, {}).get("total", 0) for w in [1, 2, 3]
    }

    windows_data = {}
    for w in [1, 2, 3]:
        w_info = grouped.get(w, {})
        windows_data[str(w)] = {
            "label":  w_info.get("label", ""),
            "total":  w_info.get("total", 0),
            "groups": w_info.get("groups", []),
        }

    return jsonify({
        "date":           date_str,
        "total_emails":   len(emails),
        "parse_failures": grouped.get("parse_failures", 0),
        "window_totals":  window_totals,
        "windows":        windows_data,
        "projects":       grouped.get("projects", []),
    })


# ─── Shift Handover ────────────────────────────────────────────────────────────

def _sh_user_name(user):
    return (user.get("full_name") or user.get("username", "")).strip()


@app.route("/shift-handover")
@login_required
def shift_handover():
    sel_date_str = request.args.get("date", date.today().isoformat())
    try:
        date.fromisoformat(sel_date_str)
    except ValueError:
        sel_date_str = date.today().isoformat()
    dashboard   = db.sh_daily_dashboard(sel_date_str)
    projects    = db.sh_get_project_names()
    daily_notes = db.sh_get_daily_notes(sel_date_str)
    stats       = db.sh_dashboard_stats(sel_date_str)
    from datetime import timedelta
    try:
        _d = date.fromisoformat(sel_date_str)
    except ValueError:
        _d = date.today()
    prev_date = (_d - timedelta(days=1)).isoformat()
    next_date = (_d + timedelta(days=1)).isoformat()
    sel_date_display = _d.strftime('%d-%m-%Y')
    return render_template("shift_handover.html",
                           app_name=APP_NAME,
                           sel_date=sel_date_str,
                           sel_date_display=sel_date_display,
                           prev_date=prev_date,
                           next_date=next_date,
                           dashboard=dashboard,
                           projects=projects,
                           daily_notes=daily_notes,
                           stats=stats,
                           current_user=g.user)


@app.route("/shift-handover/daily-notes", methods=["POST"])
@login_required
def sh_save_daily_notes():
    if not g.user or not g.user.get("sh_is_admin"):
        return jsonify({"error": "Admin only"}), 403
    note_date = request.form.get("note_date", date.today().isoformat())
    db.sh_upsert_daily_notes(
        note_date=note_date,
        duty_manager=request.form.get("duty_manager", ""),
        week_label=request.form.get("week_label", ""),
        key_issues=request.form.get("key_issues", ""),
        actions_for_tomorrow=request.form.get("actions_for_tomorrow", ""),
    )
    flash("Daily notes saved.", "success")
    return redirect(url_for("shift_handover", date=note_date))


@app.route("/shift-handover/form")
@login_required
def sh_form():
    project_name = request.args.get("project", "").strip()
    date_str     = request.args.get("date", date.today().isoformat())
    shift_num    = int(request.args.get("shift", 1))
    if not project_name:
        flash("Project is required.", "danger")
        return redirect(url_for("shift_handover"))
    try:
        date.fromisoformat(date_str)
    except ValueError:
        date_str = date.today().isoformat()

    user      = g.user
    user_name = _sh_user_name(user)
    handover  = db.sh_get_or_create_handover(date_str, shift_num, project_name,
                                              user["id"], user_name)
    handover_data = db.sh_get_handover(handover["id"])
    raw_entries   = handover_data["client_entries"]
    moms          = handover_data["moms"]

    known_clients = db.sh_get_active_clients(project_name)

    # Merge: one row per active client; fill with saved data where it exists
    _blank_entry = {
        "client_name": "", "tickets": "",
        "entry_status": "" if project_name == "Content" else "N/A",
        "engineer_worked": "", "issues": "", "engineer_notes": "",
        "manager_notes": "", "next_shift_engineer": "",
        "migration_report_sent": False, "drive_changes_alerts": False, "row_tint": None,
    }
    entries_map = {(e.get("client_name") or "").strip().lower(): e for e in raw_entries}
    entries = []
    seen_keys = set()
    for c in known_clients:
        key = c.strip().lower()
        seen_keys.add(key)
        row = dict(_blank_entry)
        row["client_name"] = c
        if key in entries_map:
            row.update(entries_map[key])
        entries.append(row)
    # Also show any saved entries for clients no longer in the active list
    for e in raw_entries:
        if (e.get("client_name") or "").strip().lower() not in seen_keys:
            entries.append(e)

    # Previous shift for reference
    prev_shift = shift_num - 1
    prev_date  = date_str
    if prev_shift < 1:
        prev_shift = 3
        from datetime import timedelta
        prev_date = (date.fromisoformat(date_str) - timedelta(days=1)).isoformat()
    prev_handover = db.sh_get_handover_by_slot(prev_date, prev_shift, project_name)
    prev_entries  = []
    if prev_handover:
        prev_entries = db.sh_get_handover(prev_handover["id"])["client_entries"]
    engineers = db.sh_get_engineers()
    curr_handler_map, next_handler_map, next_shift_label = db.sh_get_client_shift_handlers(date_str, shift_num, project_name)
    curr_shift_engineers = list(dict.fromkeys(curr_handler_map.values()))
    next_shift_engineers = list(dict.fromkeys(next_handler_map.values()))
    # Build per-client mom map for quick lookup
    mom_by_client = {}
    for m in moms:
        key = (m.get("client_name") or "").strip().lower()
        mom_by_client.setdefault(key, []).append(m)

    shift_labels = {1: "Morning (5:00 AM – 2:00 PM)", 2: "Afternoon (1:00 PM – 10:00 PM)", 3: "Night (9:00 PM – 6:00 AM)"}
    shift_times  = {1: "5:00 AM – 2:00 PM", 2: "1:00 PM – 10:00 PM", 3: "9:00 PM – 6:00 AM"}

    _d = date.fromisoformat(date_str)
    date_display = _d.strftime('%A, %B ') + str(_d.day) + _d.strftime(', %Y')

    return render_template("shift_handover_form.html",
                           app_name=APP_NAME,
                           handover=handover,
                           entries=entries,
                           known_clients=known_clients,
                           engineers=engineers,
                           curr_handler_map=curr_handler_map,
                           next_handler_map=next_handler_map,
                           curr_shift_engineers=curr_shift_engineers,
                           next_shift_engineers=next_shift_engineers,
                           next_shift_label=next_shift_label,
                           prev_handover=prev_handover,
                           prev_entries=prev_entries,
                           moms=moms,
                           mom_by_client=mom_by_client,
                           project_name=project_name,
                           date_str=date_str,
                           date_display=date_display,
                           shift_num=shift_num,
                           shift_label=shift_labels.get(shift_num, f"Shift {shift_num}"),
                           shift_time=shift_times.get(shift_num, ""),
                           current_user=user,
                           is_admin=bool(user.get("sh_is_admin")))


@app.route("/shift-handover/form/save", methods=["POST"])
@login_required
def sh_form_save():
    handover_id  = int(request.form.get("handover_id", 0))
    project_name = request.form.get("project_name", "")
    date_str     = request.form.get("date_str", date.today().isoformat())
    shift_num    = int(request.form.get("shift_num", 1))
    lead_notes   = request.form.get("lead_notes", "")
    user         = g.user
    user_name    = _sh_user_name(user)

    # Collect client entry rows from form
    client_names         = request.form.getlist("client_name[]")
    tickets_list         = request.form.getlist("tickets[]")
    statuses_list        = request.form.getlist("entry_status[]")
    eng_worked_list      = request.form.getlist("engineer_worked[]")
    issues_list          = request.form.getlist("issues[]")
    eng_notes_list       = request.form.getlist("engineer_notes[]")
    mgr_notes_list       = request.form.getlist("manager_notes[]")
    next_eng_list        = request.form.getlist("next_shift_engineer[]")
    mig_report_list      = request.form.getlist("migration_report_sent[]")
    drive_alerts_list    = request.form.getlist("drive_changes_alerts[]")
    row_tint_list        = request.form.getlist("row_tint[]")

    checked_mig  = set(request.form.getlist("migration_report_sent_check"))
    checked_drv  = set(request.form.getlist("drive_changes_alerts_check"))

    entries = []
    for i, cn in enumerate(client_names):
        cn = cn.strip()
        if not cn:
            continue
        entries.append({
            "client_name":           cn,
            "tickets":               tickets_list[i]       if i < len(tickets_list)       else "",
            "entry_status":          statuses_list[i]      if i < len(statuses_list)      else "NA",
            "engineer_worked":       eng_worked_list[i]    if i < len(eng_worked_list)    else "",
            "issues":                issues_list[i]        if i < len(issues_list)        else "",
            "engineer_notes":        eng_notes_list[i]     if i < len(eng_notes_list)     else "",
            "manager_notes":         mgr_notes_list[i]     if i < len(mgr_notes_list)     else "",
            "next_shift_engineer":   next_eng_list[i]      if i < len(next_eng_list)      else "",
            "migration_report_sent": str(i) in checked_mig,
            "drive_changes_alerts":  str(i) in checked_drv,
            "row_tint":              row_tint_list[i]      if i < len(row_tint_list)      else None,
        })

    db.sh_save_entries(handover_id, entries, user_name, lead_notes=lead_notes)
    flash("Draft saved.", "success")
    return redirect(url_for("sh_form", project=project_name, date=date_str, shift=shift_num))


@app.route("/shift-handover/form/submit", methods=["POST"])
@login_required
def sh_form_submit():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("Only shift handover admins can submit a handover.", "danger")
        return redirect(url_for("shift_handover"))
    handover_id  = int(request.form.get("handover_id", 0))
    project_name = request.form.get("project_name", "")
    date_str     = request.form.get("date_str", date.today().isoformat())
    shift_num    = int(request.form.get("shift_num", 1))
    user         = g.user
    user_name    = _sh_user_name(user)
    db.sh_submit_handover(handover_id, user["id"], user_name)
    flash("Handover submitted successfully.", "success")
    return redirect(url_for("sh_form", project=project_name, date=date_str, shift=shift_num))


@app.route("/shift-handover/<int:handover_id>/engineer-ack", methods=["POST"])
@login_required
def sh_engineer_ack(handover_id):
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin only.", "danger")
        return redirect(url_for("shift_handover"))
    db.sh_engineer_ack(handover_id, _sh_user_name(g.user))
    flash("Engineer acknowledgement recorded.", "success")
    back = request.referrer or url_for("shift_handover")
    return redirect(back)


@app.route("/shift-handover/<int:handover_id>/manager-ack", methods=["POST"])
@login_required
def sh_manager_ack(handover_id):
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin only.", "danger")
        return redirect(url_for("shift_handover"))
    db.sh_manager_ack(handover_id, _sh_user_name(g.user))
    flash("Manager acknowledgement recorded.", "success")
    back = request.referrer or url_for("shift_handover")
    return redirect(back)


@app.route("/shift-handover/<int:handover_id>/mom", methods=["POST"])
@login_required
def sh_upload_mom(handover_id):
    f = request.files.get("mom_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return redirect(request.referrer or url_for("shift_handover"))
    client_name = request.form.get("mom_client", "")
    notes       = request.form.get("mom_notes", "")
    user        = g.user
    db.sh_upload_mom(handover_id, client_name, f.filename, f.read(),
                     notes, user["id"], _sh_user_name(user))
    flash("MOM uploaded.", "success")
    return redirect(request.referrer or url_for("shift_handover"))


@app.route("/shift-handover/mom/<int:mom_id>/download")
@login_required
def sh_download_mom(mom_id):
    file_name, file_data = db.sh_get_mom_file(mom_id)
    if not file_data:
        flash("File not found.", "danger")
        return redirect(url_for("shift_handover"))
    return send_file(BytesIO(file_data), as_attachment=True, download_name=file_name)


@app.route("/shift-handover/mom/<int:mom_id>/delete", methods=["POST"])
@login_required
def sh_delete_mom(mom_id):
    user = g.user
    if not user or not user.get("sh_is_admin"):
        flash("Only SH admins can delete MOM files.", "danger")
        return redirect(request.referrer or url_for("shift_handover"))
    db.sh_delete_mom(mom_id, user["id"], True)
    flash("MOM file deleted.", "success")
    return redirect(request.referrer or url_for("shift_handover"))


@app.route("/shift-handover/history")
@login_required
def sh_history():
    page = max(1, int(request.args.get("page", 1)))
    per_page = 20
    filter_project   = request.args.get("project", "")
    filter_shift     = request.args.get("shift", "")
    filter_status    = request.args.get("status", "")
    filter_date_from = request.args.get("date_from", "")
    filter_date_to   = request.args.get("date_to", "")
    handovers, total = db.sh_get_handovers(
        project_name=filter_project or None,
        shift_num=filter_shift or None,
        status=filter_status or None,
        date_from=filter_date_from or None,
        date_to=filter_date_to or None,
        limit=per_page,
        offset=(page - 1) * per_page,
    )
    projects    = db.sh_get_project_names()
    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template("shift_handover_history.html",
                           app_name=APP_NAME,
                           handovers=handovers,
                           projects=projects,
                           total=total,
                           page=page,
                           total_pages=total_pages,
                           filter_project=filter_project,
                           filter_shift=filter_shift,
                           filter_status=filter_status,
                           filter_date_from=filter_date_from,
                           filter_date_to=filter_date_to,
                           current_user=g.user)


@app.route("/shift-handover/client-history")
@login_required
def sh_client_history():
    from collections import OrderedDict
    project_name = request.args.get("project", "")
    client_name  = request.args.get("client", "")
    from_date    = request.args.get("from_date", "")
    to_date      = request.args.get("to_date", "")

    if not project_name or not client_name:
        flash("Project and client name are required.", "warning")
        return redirect(url_for("shift_handover"))

    rows = db.sh_get_client_history(
        project_name, client_name,
        date_from=from_date or None,
        date_to=to_date or None,
    )

    shift_names = {1: "Morning", 2: "Afternoon", 3: "Night"}

    # Group by date → list of {shift_num, submitted_by_name, status, entry}
    by_date = OrderedDict()
    for r in rows:
        d = r["handover_date"]
        if d not in by_date:
            by_date[d] = []
        r["shift_name"] = shift_names.get(r["shift_num"], "")
        by_date[d].append(r)

    # Format date labels
    dates_display = {}
    for d in by_date:
        if hasattr(d, "strftime"):
            dates_display[d] = d.strftime("%A, %B %-d, %Y") if os.name != "nt" else d.strftime("%A, %B %d, %Y").replace(" 0", " ")
        else:
            dates_display[d] = str(d)

    is_content = (project_name == "Content")
    is_admin   = bool(g.user and g.user.get("sh_is_admin"))

    return render_template(
        "shift_handover_client_history.html",
        app_name=APP_NAME,
        project_name=project_name,
        client_name=client_name,
        from_date=from_date,
        to_date=to_date,
        by_date=by_date,
        dates_display=dates_display,
        is_content=is_content,
        is_admin=is_admin,
        current_user=g.user,
    )


@app.route("/shift-handover/compliance")
@login_required
def sh_compliance():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("shift_handover"))
    target_date = request.args.get("date", date.today().isoformat())
    rows = db.sh_compliance_data(target_date)
    return render_template("shift_handover_compliance.html",
                           app_name=APP_NAME,
                           rows=rows,
                           target_date=target_date,
                           current_user=g.user)


@app.route("/shift-handover/selector")
@login_required
def sh_selector():
    projects = db.sh_get_project_names()
    return render_template("shift_handover_selector.html",
                           app_name=APP_NAME,
                           projects=projects,
                           today=date.today().isoformat(),
                           current_user=g.user)


# ── Clients management ────────────────────────────────────────────────────────

@app.route("/shift-handover/clients")
@login_required
def sh_clients():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("shift_handover"))
    projects = db.sh_get_project_names()
    active_project = request.args.get("project", projects[0] if projects else "")
    clients = db.sh_get_clients(active_project) if active_project else []
    return render_template("shift_handover_clients.html",
                           app_name=APP_NAME,
                           projects=projects,
                           active_project=active_project,
                           clients=clients,
                           current_user=g.user)


@app.route("/shift-handover/projects/add", methods=["POST"])
@login_required
def sh_projects_add():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("sh_clients"))
    pname = request.form.get("project_name", "").strip()
    if pname:
        db.sh_add_project(pname)
        flash(f"Project '{pname}' added.", "success")
        return redirect(url_for("sh_clients", project=pname))
    flash("Project name is required.", "danger")
    return redirect(url_for("sh_clients"))


@app.route("/shift-handover/projects/delete", methods=["POST"])
@login_required
def sh_projects_delete():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("sh_clients"))
    pname = request.form.get("project_name", "").strip()
    if pname:
        db.sh_delete_project(pname)
        flash(f"Project '{pname}' deleted.", "success")
    return redirect(url_for("sh_clients"))


@app.route("/shift-handover/clients/add", methods=["POST"])
@login_required
def sh_clients_add():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("sh_clients"))
    project_name = request.form.get("project_name", "").strip()
    client_name  = request.form.get("client_name", "").strip()
    if project_name and client_name:
        db.sh_add_client(project_name, client_name)
        flash(f"Client '{client_name}' added to {project_name}.", "success")
    else:
        flash("Project and client name are required.", "danger")
    return redirect(url_for("sh_clients", project=project_name))


@app.route("/shift-handover/clients/<int:client_id>/toggle", methods=["POST"])
@login_required
def sh_clients_toggle(client_id):
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("sh_clients"))
    db.sh_toggle_client(client_id)
    back_project = request.form.get("project_name", "")
    return redirect(url_for("sh_clients", project=back_project))


@app.route("/shift-handover/clients/<int:client_id>/delete", methods=["POST"])
@login_required
def sh_clients_delete(client_id):
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("sh_clients"))
    db.sh_delete_client(client_id)
    flash("Client deleted.", "success")
    back_project = request.form.get("project_name", "")
    return redirect(url_for("sh_clients", project=back_project))


# ── Users management ─────────────────────────────────────────────────────────

@app.route("/shift-handover/users")
@login_required
def sh_users():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("shift_handover"))
    users = db.sh_get_all_users_with_shifts()
    return render_template("shift_handover_users.html",
                           app_name=APP_NAME,
                           users=users,
                           current_user=g.user)


@app.route("/shift-handover/users/<int:user_id>/role", methods=["POST"])
@login_required
def sh_users_update_role(user_id):
    if not g.user or not g.user.get("sh_is_admin"):
        return jsonify({"error": "SH Admin only"}), 403
    new_role = request.form.get("role", "user")
    if new_role not in ("admin", "user"):
        return jsonify({"error": "Invalid role"}), 400
    db.sh_set_user_sh_admin(user_id, new_role == "admin")
    flash("SH Admin access updated.", "success")
    return redirect(url_for("sh_users"))


@app.route("/shift-handover/users/<int:user_id>/shifts", methods=["POST"])
@login_required
def sh_users_update_shifts(user_id):
    if not g.user or not g.user.get("sh_is_admin"):
        return jsonify({"error": "SH Admin only"}), 403
    s1 = "shift_1" in request.form
    s2 = "shift_2" in request.form
    s3 = "shift_3" in request.form
    db.sh_update_user_shifts(user_id, s1, s2, s3)
    flash("Shift assignments updated.", "success")
    return redirect(url_for("sh_users"))


@app.route("/shift-handover/users/<int:user_id>/toggle", methods=["POST"])
@login_required
def sh_users_toggle(user_id):
    if not g.user or not g.user.get("sh_is_admin"):
        return jsonify({"error": "SH Admin only"}), 403
    db.sh_toggle_user_status(user_id)
    flash("User status updated.", "success")
    return redirect(url_for("sh_users"))


@app.route("/shift-handover/users/<int:user_id>/delete", methods=["POST"])
@login_required
def sh_users_delete(user_id):
    if not g.user or not g.user.get("sh_is_admin"):
        return jsonify({"error": "SH Admin only"}), 403
    if user_id == g.user.get("id"):
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("sh_users"))
    db.sh_delete_user(user_id)
    flash("User deleted.", "success")
    return redirect(url_for("sh_users"))


# ── Reports ────────────────────────────────────────────────────────────────────

@app.route("/shift-handover/reports")
@login_required
def sh_reports():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("shift_handover"))
    mode      = request.args.get("mode", "date")
    date_val  = request.args.get("date", date.today().isoformat())
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    q         = request.args.get("q", "").strip()
    rows = []
    if mode == "date" and date_val:
        rows = db.sh_report_by_date(date_val)
    elif mode == "employee" and (q or date_from or date_to):
        rows = db.sh_report_by_employee(q, date_from or None, date_to or None)
    elif mode == "client" and (q or date_from or date_to):
        rows = db.sh_report_by_client(q, date_from or None, date_to or None)
    projects = db.sh_get_project_names()
    return render_template("shift_handover_reports.html",
                           app_name=APP_NAME,
                           mode=mode,
                           date_val=date_val,
                           date_from=date_from,
                           date_to=date_to,
                           q=q,
                           rows=rows,
                           projects=projects,
                           current_user=g.user)


# ── History Excel export ───────────────────────────────────────────────────────

@app.route("/shift-handover/history/excel")
@login_required
def sh_history_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    filter_project   = request.args.get("project", "")
    filter_shift     = request.args.get("shift", "")
    filter_status    = request.args.get("status", "")
    filter_date_from = request.args.get("date_from", "")
    filter_date_to   = request.args.get("date_to", "")
    rows = db.sh_get_history_for_excel(
        project_name=filter_project or None,
        shift_num=filter_shift or None,
        status=filter_status or None,
        date_from=filter_date_from or None,
        date_to=filter_date_to or None,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Handover History"
    headers = ["Date", "Shift", "Project", "Submitted By", "Status", "Lead Notes",
               "Eng Ack", "Mgr Ack", "Client", "Tickets", "Entry Status",
               "Engineer Worked", "Issues", "Engineer Notes", "Next Shift Engineer"]
    hdr_fill = PatternFill("solid", fgColor="1a3a5c")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            str(row.get("handover_date", "")),
            f"Shift {row.get('shift_num', '')}",
            row.get("project_name", ""),
            row.get("submitted_by_name", ""),
            row.get("status", ""),
            row.get("lead_notes", ""),
            "Yes" if row.get("engineer_acknowledged") else "No",
            "Yes" if row.get("manager_acknowledged") else "No",
            row.get("client_name", ""),
            row.get("tickets", ""),
            row.get("entry_status", ""),
            row.get("engineer_worked", ""),
            row.get("issues", ""),
            row.get("engineer_notes", ""),
            row.get("next_shift_engineer", ""),
        ])
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"handover_history_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Tracking (enhanced compliance) ───────────────────────────────────────────

@app.route("/shift-handover/tracking")
@login_required
def sh_tracking():
    if not g.user or not g.user.get("sh_is_admin"):
        flash("SH Admin access required.", "danger")
        return redirect(url_for("shift_handover"))
    from datetime import timedelta
    target_date = request.args.get("date", date.today().isoformat())
    try:
        d = date.fromisoformat(target_date)
    except ValueError:
        d = date.today()
        target_date = d.isoformat()
    prev_date = (d - timedelta(days=1)).isoformat()
    next_date = (d + timedelta(days=1)).isoformat()
    data = db.sh_engineer_activity(target_date)
    return render_template("shift_handover_tracking.html",
                           app_name=APP_NAME,
                           target_date=target_date,
                           prev_date=prev_date,
                           next_date=next_date,
                           data=data,
                           current_user=g.user)


if __name__ == "__main__":
    app.run(debug=True, port=5050)
