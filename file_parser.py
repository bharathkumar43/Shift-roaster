import os
import re
from io import BytesIO

from openpyxl import load_workbook

TESSERACT_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.getenv("USERNAME", "")),
]

DAY_MAP = {
    "mon": "Monday", "monday": "Monday",
    "tue": "Tuesday", "tues": "Tuesday", "tuesday": "Tuesday",
    "wed": "Wednesday", "wednesday": "Wednesday",
    "thu": "Thursday", "thur": "Thursday", "thurs": "Thursday", "thursday": "Thursday",
    "fri": "Friday", "friday": "Friday",
    "sat": "Saturday", "saturday": "Saturday",
    "sun": "Sunday", "sunday": "Sunday",
}

VALID_TYPES = {"content", "email", "message"}


def parse_file(file_storage):
    """
    Parse an uploaded file and return a list of employee dicts.
    Supports .xlsx, .pdf, .png, .jpg, .jpeg.

    Each returned dict: {name, content_types: [], working_days: [], shift: int|None}
    Raises ValueError on parse failure.
    """
    filename = file_storage.filename.lower()
    file_bytes = file_storage.read()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return parse_excel(file_bytes)
    elif filename.endswith(".pdf"):
        return parse_pdf(file_bytes)
    elif filename.endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff")):
        return parse_image(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


def parse_excel(file_bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()

    if not rows:
        raise ValueError("Excel file is empty.")

    return _extract_names(rows)


def parse_pdf(file_bytes):
    try:
        import pdfplumber
    except ImportError:
        raise ValueError("pdfplumber is not installed. Run: pip install pdfplumber")

    rows = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        cleaned = [str(c).strip() if c else "" for c in row]
                        rows.append(cleaned)
            else:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        parts = re.split(r"\s{2,}|\t", line.strip())
                        if len(parts) >= 3:
                            rows.append(parts)

    if not rows:
        raise ValueError("Could not extract any data from the PDF.")

    return _extract_names(rows)


def parse_image(file_bytes):
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        raise ValueError("pytesseract and Pillow are required. Run: pip install pytesseract Pillow")

    for path in TESSERACT_PATHS:
        if os.path.isfile(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break

    img = Image.open(BytesIO(file_bytes))
    text = pytesseract.image_to_string(img)

    if not text or not text.strip():
        raise ValueError("Could not extract any text from the image. Ensure Tesseract OCR is installed.")

    rows = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = re.split(r"\s{2,}|\t|\|", line)
        parts = [p.strip() for p in parts if p.strip()]
        if parts:
            rows.append(parts)

    if not rows:
        raise ValueError("Could not extract any text rows from the image.")

    return _extract_names(rows)


def _extract_names(rows):
    """
    Extract employee names from raw rows.
    Skips the first row if it looks like a header.
    For each remaining row, takes the first cell that contains letters as the name.
    Returns a list of employee dicts with default working_days and content_types.
    """
    _SKIP_VALS = {'none', 'nan', 'null', 'n/a', '-', ''}
    _HEADER_HINTS = {'name', 'employee', 'staff', 'person', 'worker',
                     'sno', 's.no', 'sl.no', '#', 'no', 'id', 'sr', 'sr.no', 'sl'}

    # Decide where data rows start (skip header row if detected)
    start = 0
    if rows:
        first_vals = [str(v).strip().lower() for v in rows[0] if str(v).strip()]
        if any(v in _HEADER_HINTS or 'name' in v or 'employee' in v for v in first_vals):
            start = 1

    employees = []
    for row in rows[start:]:
        name = ""
        project_names = []
        name_found = False

        for cell in row:
            val = str(cell).strip() if cell is not None else ""
            if not val or val.lower() in _SKIP_VALS:
                continue
            if re.match(r'^[\d\s.\-/#+@()]+$', val):
                continue

            if not name_found:
                name = val
                name_found = True
            else:
                # Any subsequent text cell is treated as project name(s)
                for p in re.split(r'[,;]+', val):
                    p = p.strip()
                    if p and p.lower() not in _SKIP_VALS:
                        project_names.append(p)

        if not name or name.lower() in _SKIP_VALS:
            continue

        employees.append({
            "name": name,
            "content_types": ["Content"],
            "working_days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
            "shift": None,
            "projects": [{"name": p, "product_type": "Content"} for p in project_names],
        })

    if not employees:
        raise ValueError("No employee names found in the file.")
    return employees


def _parse_table(rows):
    """
    Given a list of row-lists, detect headers, map columns, and parse employees.
    """
    header_idx, col_map = _detect_headers(rows)
    if col_map is None:
        raise ValueError(
            "Could not detect required columns. "
            "Expected columns containing: Name, Shift, Working Days/Days, Product Type/Type."
        )

    employees = []
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if len(row) <= max(col_map.values()):
            row = row + [""] * (max(col_map.values()) + 1 - len(row))

        name = row[col_map["name"]].strip()
        if not name or name.lower() in ("none", "nan", ""):
            continue

        shift = _parse_shift(row[col_map["shift"]]) if "shift" in col_map else None
        working_days = _parse_days(row[col_map["days"]]) if "days" in col_map else []
        content_types = _parse_types(row[col_map["type"]]) if "type" in col_map else []

        if not content_types:
            content_types = ["Content"]
        if not working_days:
            working_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

        employees.append({
            "name": name,
            "content_types": content_types,
            "working_days": working_days,
            "shift": shift,
        })

    if not employees:
        raise ValueError("No valid employee rows found in the file.")

    return employees


def _detect_headers(rows):
    """Find the header row and map column indices."""
    for idx, row in enumerate(rows):
        lower = [str(c).lower().strip() for c in row]
        col_map = {}

        for i, val in enumerate(lower):
            if not val:
                continue
            if any(k in val for k in ["name", "employee"]):
                col_map["name"] = i
            elif any(k in val for k in ["shift"]):
                col_map["shift"] = i
            elif any(k in val for k in ["day", "working", "schedule"]):
                col_map["days"] = i
            elif any(k in val for k in ["type", "product", "content", "category"]):
                col_map["type"] = i

        if "name" in col_map and len(col_map) >= 2:
            return idx, col_map

    return 0, None


def _parse_shift(val):
    val = str(val).strip().lower()
    match = re.search(r"(\d)", val)
    if match:
        s = int(match.group(1))
        if 1 <= s <= 3:
            return s
    return None


def _parse_days(val):
    val = str(val).strip()
    if not val:
        return []

    days = []
    parts = re.split(r"[,;/\s]+", val.lower())
    for p in parts:
        p = p.strip().rstrip(".")
        if p in DAY_MAP:
            full = DAY_MAP[p]
            if full not in days:
                days.append(full)

    return days


def _parse_types(val):
    val = str(val).strip()
    if not val:
        return []

    types = []
    parts = re.split(r"[,;/&]+", val.lower())
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if "content" in p or p in ("con", "cnt"):
            if "Content" not in types:
                types.append("Content")
        elif "email" in p or p in ("eml", "e-mail"):
            if "Email" not in types:
                types.append("Email")
        elif "message" in p or p in ("msg",):
            if "Message" not in types:
                types.append("Message")

    return types


def get_excel_headers_info(file_bytes):
    """
    Returns (headers, sample_rows, auto_col_map) for the column mapping UI.
    auto_col_map maps field names ('name','shift','days','type') to column indices.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
        if len(all_rows) >= 6:
            break
    wb.close()

    if not all_rows:
        return [], [], {}

    headers = all_rows[0]
    sample_rows = all_rows[1:]
    _, auto_map = _detect_headers(all_rows)
    return headers, sample_rows, auto_map or {}


def parse_excel_with_col_map(file_bytes, col_map):
    """
    Parse Excel using an explicit column mapping dict.
    col_map keys: 'name' (required), 'shift', 'days', 'type' (all optional).
    Values are integer column indices.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()

    if len(rows) < 2:
        raise ValueError("Excel file has no data rows.")

    employees = []
    for i in range(1, len(rows)):
        row = rows[i]
        valid_indices = [v for v in col_map.values() if v is not None]
        if valid_indices:
            max_idx = max(valid_indices)
            if len(row) <= max_idx:
                row = row + [""] * (max_idx + 1 - len(row))

        name_idx = col_map.get("name")
        if name_idx is None:
            continue
        name = row[name_idx].strip()
        if not name or name.lower() in ("none", "nan", ""):
            continue

        shift = _parse_shift(row[col_map["shift"]]) if col_map.get("shift") is not None else None
        working_days = _parse_days(row[col_map["days"]]) if col_map.get("days") is not None else []
        content_types = _parse_types(row[col_map["type"]]) if col_map.get("type") is not None else []

        if not content_types:
            content_types = ["Content"]
        if not working_days:
            working_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

        employees.append({
            "name": name,
            "content_types": content_types,
            "working_days": working_days,
            "shift": shift,
        })

    if not employees:
        raise ValueError("No valid employee rows found in the file.")
    return employees
